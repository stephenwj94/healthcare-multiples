"""
Excel override reader — loads FactSet override values from the Excel template.

Override logic:
  - Reads all segment sheets from factset_overrides.xlsx
  - Returns a dict keyed by ticker with non-blank field values
  - Blank cells are skipped (no override for that field)
  - The caller merges these over the DB snapshot data

History sheets:
  - Reads weekly multiples history for time-series charts
  - 13 sheets: NTM/LTM Rev/GP/EBITDA x, GA Rev/GP, Price, EV, Growth, Margins
  - Returns per-company rows with date, segment, and metric values

Usage:
    from fetcher.excel_override import load_overrides, load_multiples_history
    overrides = load_overrides()  # {ticker: {db_field: value, ...}, ...}
    history = load_multiples_history()  # [{date, segment, ticker, ntm_tev_rev, ...}, ...]
"""

import logging
from pathlib import Path
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

# Map from Excel header → DB field name
# Must match COLUMN_SPEC in create_excel_template.py
HEADER_TO_DB_FIELD = {
    # ── Identifiers ──
    "Ticker":               "ticker",
    "Company":              "name",
    "Segment":              "segment",
    "Sub-Segment":          "sub_segment",

    # ── Market Data ──
    "Market Cap":           "market_cap",
    "Enterprise Value":     "enterprise_value",
    "Stock Price":          "current_price",
    "52-Week High":         "fifty_two_week_high",
    "Shares Outstanding":   "shares_outstanding",
    "Total Debt":           "total_debt",
    "Total Cash":           "total_cash",
    "FY End Month":         "fy_end_month",

    # ── LTM Financials ──
    "LTM Revenue":          "ltm_revenue",
    "LTM Gross Profit":     "ltm_gross_profit",
    "LTM EBITDA":           "ltm_ebitda",
    "LTM Operating Income": "ltm_operating_income",
    "LTM Net Income":       "ltm_net_income",
    "LTM Free Cash Flow":   "ltm_fcf",
    "LTM R&D Expense":      "ltm_rd_expense",
    "LTM S&M Expense":      "ltm_sm_expense",
    "LTM G&A Expense":      "ltm_ga_expense",
    "LTM SBC":              "ltm_sbc",
    "LTM CapEx":            "ltm_capex",

    # ── Margins ──
    "Gross Margin":         "gross_margin",
    "EBITDA Margin":        "ebitda_margin",
    "Operating Margin":     "operating_margin",
    "Net Margin":           "net_margin",
    "FCF Margin":           "fcf_margin",
    "R&D % Revenue":        "rd_pct_revenue",
    "S&M % Revenue":        "sm_pct_revenue",
    "SBC % Revenue":        "sbc_pct_revenue",

    # ── Consensus Estimates ──
    "Cur FY Rev Est":       "current_fy_rev_est",
    "Next FY Rev Est":      "next_fy_rev_est",
    "Cur FY GP Est":        "current_fy_gp_est",
    "Next FY GP Est":       "next_fy_gp_est",
    "Cur FY EBITDA Est":    "current_fy_ebitda_est",
    "Next FY EBITDA Est":   "next_fy_ebitda_est",
    "Cur FY FCF Est":       "current_fy_fcf_est",
    "Next FY FCF Est":      "next_fy_fcf_est",
    "NTM Revenue":          "ntm_revenue",
    "NTM Gross Profit":     "ntm_gross_profit",
    "NTM EBITDA":           "ntm_ebitda",
    "NTM FCF":              "ntm_fcf",

    # ── Growth ──
    "NTM Revenue Growth":   "ntm_revenue_growth",
    "Cur FY Rev Growth":    "current_fy_rev_growth",
    "Next FY Rev Growth":   "next_fy_rev_growth",
    "3Y Rev CAGR":          "n3y_revenue_cagr",
    "5Y Growth Rate":       "five_year_growth_rate",
    "LTM Rev Growth YoY":   "ltm_revenue_growth",
    "Net Revenue Retention": "net_revenue_retention",

    # ── Price Performance ──
    "2-Week Price Chg":     "price_change_2w",
    "2-Month Price Chg":    "price_change_2m",
    "6-Month Price Chg":    "price_change_6m",
    "1-Year Price Chg":     "price_change_1y",
    "YTD Price Chg":        "price_change_ytd",

    # ── Legacy headers (backward compatibility with old templates) ──
    "Current FY Rev Est":   "current_fy_rev_est",
    "Next FY Rev Est":      "next_fy_rev_est",
    "Current FY EBITDA Est":"current_fy_ebitda_est",
    "Next FY EBITDA Est":   "next_fy_ebitda_est",
}

# Fields that should NOT be overridden (identifier columns)
ID_FIELDS = {"ticker", "name", "segment", "sub_segment"}

# ── Unit conversion: FactSet Excel uses $M, DB stores raw dollars ──
# Fields that come from FactSet in MILLIONS and need × 1,000,000
MILLIONS_FIELDS = {
    # Market data
    "market_cap",
    "enterprise_value",
    "total_debt",
    "total_cash",
    "shares_outstanding",          # FactSet reports in millions

    # LTM Financials
    "ltm_revenue",
    "ltm_gross_profit",
    "ltm_ebitda",
    "ltm_operating_income",
    "ltm_net_income",
    "ltm_fcf",
    "ltm_rd_expense",
    "ltm_sm_expense",
    "ltm_ga_expense",
    "ltm_sbc",
    "ltm_capex",

    # Consensus Estimates
    "current_fy_rev_est",
    "next_fy_rev_est",
    "current_fy_gp_est",
    "next_fy_gp_est",
    "current_fy_ebitda_est",
    "next_fy_ebitda_est",
    "current_fy_fcf_est",
    "next_fy_fcf_est",
    "ntm_revenue",
    "ntm_gross_profit",
    "ntm_ebitda",
    "ntm_fcf",
}

# Fields already in correct units (no conversion needed):
#   current_price, fifty_two_week_high  → raw $
#   fy_end_month                        → integer 1-12
#   all margin fields                   → decimal (0.75 = 75%)
#   all growth fields                   → decimal (0.20 = 20%)
#   all price change fields             → decimal (-0.05 = -5%)

# Fields that are integers (not floats)
INTEGER_FIELDS = {"fy_end_month"}

# Fields where the app should recalculate derived values after override
TRIGGERS_RECALC = {
    # Core valuation inputs
    "enterprise_value", "market_cap",
    "ntm_revenue", "ntm_ebitda", "ntm_gross_profit", "ntm_fcf",
    "ltm_revenue", "ltm_gross_profit", "ltm_ebitda",
    "gross_margin", "ebitda_margin",
    "ntm_revenue_growth",
    "current_fy_rev_est", "next_fy_rev_est",
    "current_fy_gp_est", "next_fy_gp_est",
    "current_fy_ebitda_est", "next_fy_ebitda_est",
    "current_fy_fcf_est", "next_fy_fcf_est",
    "fifty_two_week_high", "current_price",
    # New financial fields that feed derived metrics
    "ltm_operating_income", "ltm_net_income", "ltm_fcf",
    "ltm_rd_expense", "ltm_sm_expense", "ltm_ga_expense",
    "ltm_sbc", "ltm_capex",
}


def load_overrides(excel_path):
    """
    Load override values from the FactSet Excel template.

    Args:
        excel_path: Path to the factset_overrides.xlsx file

    Returns:
        dict: {ticker: {db_field: value, ...}} for all non-blank override cells.
              Only includes fields that have actual values entered.
              Returns empty dict if file doesn't exist or has errors.
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        logger.info(f"No override file found at {excel_path}")
        return {}

    # Only read our known segment override sheets
    OVERRIDE_SHEETS = {"Horizontal SW", "Vertical SW", "Infrastructure", "Cybersecurity"}

    try:
        overrides = {}
        xls = pd.ExcelFile(excel_path, engine="openpyxl")

        for sheet_name in xls.sheet_names:
            if sheet_name not in OVERRIDE_SHEETS:
                continue

            # Row 0 = group labels (skip), Row 1 = column headers, Row 2 = unit hints (skip)
            # Data starts at row 3 (0-indexed after header)
            df = pd.read_excel(
                xls, sheet_name=sheet_name,
                header=1,      # Column headers are in row 2 (0-indexed row 1)
                skiprows=[2],  # Skip the unit-hint row (row 3 in Excel = index 2 after header)
                engine="openpyxl"
            )

            if "Ticker" not in df.columns:
                logger.warning(f"Sheet '{sheet_name}' missing Ticker column, skipping")
                continue

            for _, row in df.iterrows():
                ticker = row.get("Ticker")
                if pd.isna(ticker) or not str(ticker).strip():
                    continue

                ticker = str(ticker).strip().upper().replace("-US", "")
                company_overrides = {}

                for excel_header, db_field in HEADER_TO_DB_FIELD.items():
                    if db_field in ID_FIELDS:
                        continue  # Don't override identifiers

                    if excel_header not in df.columns:
                        continue

                    val = row.get(excel_header)

                    # Skip blank/NaN cells — these mean "no override"
                    if pd.isna(val):
                        continue

                    # Convert to appropriate Python type
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        continue  # Skip non-numeric values for numeric fields

                    # Integer fields
                    if db_field in INTEGER_FIELDS:
                        val = int(val)
                    # Unit conversion: FactSet $M → raw dollars
                    elif db_field in MILLIONS_FIELDS:
                        val = val * 1_000_000

                    company_overrides[db_field] = val

                if company_overrides:
                    overrides[ticker] = company_overrides

        if overrides:
            total_fields = sum(len(v) for v in overrides.values())
            logger.info(
                f"Loaded {total_fields} override values for "
                f"{len(overrides)} companies from {excel_path.name}"
            )
        else:
            logger.info(f"Override file exists but no values entered yet")

        return overrides

    except Exception as e:
        logger.error(f"Error reading override file {excel_path}: {e}")
        return {}


def apply_overrides(snapshot_data, overrides, recalculate=True, skip_sources=None):
    """
    Merge Excel overrides into snapshot data (list of dicts from DB).

    Args:
        snapshot_data: list of dicts from db_manager.get_latest_snapshots()
        overrides: dict from load_overrides()
        recalculate: if True, recalculate derived fields after override
        skip_sources: set of data_source values to skip (e.g. {"factset"})

    Returns:
        list of dicts with override values merged in
    """
    if not overrides:
        return snapshot_data

    result = []
    override_count = 0

    for record in snapshot_data:
        record = dict(record)  # Make a copy
        ticker = record.get("ticker", "").upper()

        # Skip overrides for records from authoritative sources
        if skip_sources and record.get("data_source", "") in skip_sources:
            result.append(record)
            continue

        if ticker in overrides:
            for field, value in overrides[ticker].items():
                record[field] = value
                override_count += 1

            if recalculate:
                record = _recalculate_derived(record)

        result.append(record)

    if override_count > 0:
        logger.info(f"Applied {override_count} overrides across {len(overrides)} companies")

    return result


def _recalculate_derived(record):
    """
    Recalculate derived fields after an override.
    This ensures multiples stay consistent with overridden fundamentals.
    """
    ev = record.get("enterprise_value")
    ntm_rev = record.get("ntm_revenue")
    ntm_ebitda = record.get("ntm_ebitda")
    ntm_gp_direct = record.get("ntm_gross_profit")
    ntm_fcf = record.get("ntm_fcf")
    ltm_rev = record.get("ltm_revenue")
    ltm_gp = record.get("ltm_gross_profit")
    ltm_ebitda = record.get("ltm_ebitda")
    ltm_opinc = record.get("ltm_operating_income")
    ltm_ni = record.get("ltm_net_income")
    ltm_fcf = record.get("ltm_fcf")
    ltm_rd = record.get("ltm_rd_expense")
    ltm_sm = record.get("ltm_sm_expense")
    ltm_ga = record.get("ltm_ga_expense")
    ltm_sbc = record.get("ltm_sbc")
    gross_margin = record.get("gross_margin")
    ntm_rev_growth = record.get("ntm_revenue_growth")
    price = record.get("current_price")
    high_52 = record.get("fifty_two_week_high")

    # ── Valuation Multiples ──

    # NTM TEV/Revenue
    if ev and ntm_rev and ev > 0 and ntm_rev > 0:
        record["ntm_tev_rev"] = ev / ntm_rev

    # NTM TEV/Gross Profit (prefer direct NTM GP if available)
    ntm_gp = ntm_gp_direct
    if not ntm_gp and ntm_rev and gross_margin and ntm_rev > 0 and gross_margin > 0:
        ntm_gp = ntm_rev * gross_margin
    if ev and ntm_gp and ev > 0 and ntm_gp > 0:
        record["ntm_tev_gp"] = ev / ntm_gp

    # NTM TEV/EBITDA
    if ev and ntm_ebitda and ev > 0 and ntm_ebitda > 0:
        record["ntm_tev_ebitda"] = ev / ntm_ebitda

    # LTM TEV/Revenue
    if ev and ltm_rev and ev > 0 and ltm_rev > 0:
        record["ltm_tev_rev"] = ev / ltm_rev

    # LTM TEV/Gross Profit
    if ev and ltm_gp and ev > 0 and ltm_gp > 0:
        record["ltm_tev_gp"] = ev / ltm_gp

    # LTM TEV/EBITDA
    if ev and ltm_ebitda and ev > 0 and ltm_ebitda > 0:
        record["ltm_tev_ebitda"] = ev / ltm_ebitda

    # ── Growth-Adjusted Multiples ──
    if ntm_rev_growth and ntm_rev_growth > 0:
        ntm_tev_rev = record.get("ntm_tev_rev")
        if ntm_tev_rev and ntm_tev_rev > 0:
            record["growth_adj_rev"] = ntm_tev_rev / (ntm_rev_growth * 100)
        ntm_tev_gp = record.get("ntm_tev_gp")
        if ntm_tev_gp and ntm_tev_gp > 0:
            record["growth_adj_gp"] = ntm_tev_gp / (ntm_rev_growth * 100)

    # ── Margins (recalc from LTM financials when available) ──
    if ltm_rev and ltm_rev > 0:
        if ltm_gp:
            record["gross_margin"] = ltm_gp / ltm_rev
        if ltm_ebitda:
            record["ebitda_margin"] = ltm_ebitda / ltm_rev
        if ltm_opinc:
            record["operating_margin"] = ltm_opinc / ltm_rev
        if ltm_ni:
            record["net_margin"] = ltm_ni / ltm_rev
        if ltm_fcf:
            record["fcf_margin"] = ltm_fcf / ltm_rev
        if ltm_rd:
            record["rd_pct_revenue"] = ltm_rd / ltm_rev
        if ltm_sm:
            record["sm_pct_revenue"] = ltm_sm / ltm_rev
        if ltm_sbc:
            record["sbc_pct_revenue"] = ltm_sbc / ltm_rev

    # ── Price ──
    # % of 52-week high
    if price and high_52 and high_52 > 0:
        record["pct_52wk_high"] = price / high_52

    return record


# ── Multiples History (weekly time-series from FactSet) ──────────────────────

# Map Excel sheet name → DB-compatible metric field name
HISTORY_SHEET_TO_FIELD = {
    # Valuation multiples
    "NTM Rev x History":    "ntm_tev_rev",
    "NTM GP x History":     "ntm_tev_gp",
    "NTM EBITDA x History": "ntm_tev_ebitda",
    "LTM Rev x History":    "ltm_tev_rev",
    "LTM GP x History":     "ltm_tev_gp",
    "LTM EBITDA x History": "ltm_tev_ebitda",
    # Growth-adjusted
    "GA Rev History":       "growth_adj_rev",
    "GA GP History":        "growth_adj_gp",
    # Fundamentals time-series
    "Stock Price History":  "current_price",
    "EV History":           "enterprise_value",
    "NTM Rev Gr History":   "ntm_revenue_growth",
    "Gross Margin History": "gross_margin",
    "EBITDA Margin History":"ebitda_margin",
}

# Map Excel segment display names → DB segment keys
EXCEL_SEGMENT_TO_KEY = {
    "Pharma":                              "pharma",
    "Consumer Health":                     "consumer_health",
    "MedTech":                             "medtech",
    "Life Sci Tools":                      "life_sci_tools",
    "Life Sci Tools / Dx / Bioprocessing": "life_sci_tools",
    "Services":                            "services",
    "Asset-Light Services":                "services",
    "CDMOs":                               "cdmo",
    "CDMO":                                "cdmo",
    "Health Tech":                         "health_tech",
}


def _read_transposed_history_sheet(wb, sheet_name, field_name):
    """
    Read a single transposed FactSet history sheet.

    New format (FactSet add-in output):
      - Row 1:  Title
      - Rows 2-6: Metadata (empty, Start Date, End Date, Currency, Frequency)
      - Row 7:  Section headers across columns (e.g. "EV", "NTM Revenue", "NTM TEV/ Revenue")
      - Row 8:  "Ticker" label + ticker symbols across columns
      - Row 9:  "Company" label + company names
      - Row 10: "Segment" label + segment display names
      - Row 11: "Date" label (header row for data)
      - Row 12+: Data rows — date in col 3 or 4, values in subsequent columns

    Multi-section sheets (3 or 2 sections): we take the LAST section
    (the computed multiple/margin/growth-adjusted value).
    Single-section sheets: use the only section.

    Returns a long-format DataFrame: [ticker, segment, date, <field_name>]
    """
    ws = wb[sheet_name]

    # Read all rows into memory at once (efficient for read_only mode)
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 12:
        return None

    # Row indices (0-based in our list, 1-based in Excel)
    row7 = all_rows[6]   # Section headers
    row8 = all_rows[7]   # Tickers
    row10 = all_rows[9]  # Segments

    # 1. Find section headers in row 7
    sections = []
    for col_idx, v in enumerate(row7):
        if v:
            sections.append(col_idx)

    if not sections:
        return None

    # Use the LAST section (computed value)
    last_section_idx = sections[-1]

    # 2. Extract tickers (row 8) and segments (row 10) from last section
    tickers = []
    segments = []
    data_col_indices = []
    for col_idx in range(last_section_idx + 1, len(row8)):
        t = row8[col_idx]
        if t and str(t).strip() and str(t).strip() != "Ticker":
            tickers.append(str(t).strip().upper().replace("-US", ""))
            seg_val = row10[col_idx] if col_idx < len(row10) else None
            segments.append(str(seg_val).strip() if seg_val else "")
            data_col_indices.append(col_idx)

    if not tickers:
        return None

    # 3. Determine date column: col 3 (idx 2) for multi-section, col 4 (idx 3) for single
    date_col_idx = None
    row12 = all_rows[11]
    for try_idx in [2, 3]:
        if try_idx < len(row12) and row12[try_idx] is not None:
            try:
                pd.Timestamp(row12[try_idx])
                date_col_idx = try_idx
                break
            except (ValueError, TypeError):
                pass

    if date_col_idx is None:
        # Fallback: date might be at last_section_idx itself
        if last_section_idx < len(row12) and row12[last_section_idx] is not None:
            try:
                pd.Timestamp(row12[last_section_idx])
                date_col_idx = last_section_idx
            except (ValueError, TypeError):
                pass

    if date_col_idx is None:
        return None

    # 4. Read all data rows (row 12+ = index 11+)
    rows = []
    for row in all_rows[11:]:
        if date_col_idx >= len(row):
            continue
        date_val = row[date_col_idx]
        if date_val is None:
            continue
        try:
            date_str = pd.Timestamp(date_val).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            continue

        for i, col_idx in enumerate(data_col_indices):
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "":
                continue
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue

            seg_key = EXCEL_SEGMENT_TO_KEY.get(segments[i], segments[i].lower().replace(" ", "_"))
            rows.append({
                "ticker": tickers[i],
                "segment": seg_key,
                "date": date_str,
                field_name: val,
            })

    if not rows:
        return None

    return pd.DataFrame(rows)


def load_multiples_history(excel_path):
    """
    Load weekly multiples history from the FactSet Excel template.

    Supports the transposed FactSet format where tickers are in columns
    and dates are in rows.  Each history sheet may have 1-3 sections;
    the LAST section contains the computed metric (multiple, margin, etc.).

    Args:
        excel_path: Path to the factset_overrides.xlsx file

    Returns:
        list of dicts: [{date, segment, ticker, ntm_tev_rev, ntm_tev_gp,
                         ntm_tev_ebitda, ltm_tev_rev, ltm_tev_gp,
                         ltm_tev_ebitda, growth_adj_rev, growth_adj_gp,
                         current_price, enterprise_value, ntm_revenue_growth,
                         gross_margin, ebitda_margin}, ...]
        Returns empty list if file doesn't exist or has errors.
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        logger.info(f"No override file found at {excel_path}")
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        available_sheets = set(wb.sheetnames)

        all_long_frames = []

        for sheet_name, field_name in HISTORY_SHEET_TO_FIELD.items():
            if sheet_name not in available_sheets:
                logger.debug(f"History sheet '{sheet_name}' not found, skipping")
                continue

            try:
                df = _read_transposed_history_sheet(wb, sheet_name, field_name)
                if df is not None and not df.empty:
                    all_long_frames.append(df)
                    logger.debug(
                        f"  {sheet_name}: {len(df)} rows, "
                        f"{df['ticker'].nunique()} companies"
                    )
            except Exception as e:
                logger.warning(f"Error reading history sheet '{sheet_name}': {e}")
                continue

        wb.close()

        if not all_long_frames:
            logger.info("No multiples history data found in Excel")
            return []

        # Merge all metrics on (ticker, segment, date)
        merged = all_long_frames[0]
        for frame in all_long_frames[1:]:
            merged = pd.merge(
                merged, frame,
                on=["ticker", "segment", "date"],
                how="outer",
            )

        # Sort by date and segment
        merged = merged.sort_values(["date", "segment", "ticker"])

        # Convert to list of dicts
        result = merged.to_dict("records")

        # Clean up NaN → None
        for row in result:
            for key, val in row.items():
                if isinstance(val, float) and np.isnan(val):
                    row[key] = None

        unique_dates = merged["date"].nunique()
        unique_tickers = merged["ticker"].nunique()
        logger.info(
            f"Loaded multiples history: {len(result)} rows, "
            f"{unique_tickers} companies, {unique_dates} weekly dates"
        )

        return result

    except Exception as e:
        logger.error(f"Error reading multiples history from {excel_path}: {e}")
        return []
