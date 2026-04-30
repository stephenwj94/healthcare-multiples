"""
Generate the FactSet override Excel template.

Run once to create data/factset_overrides.xlsx with:
  - One sheet per segment (Pharma, Consumer Health, MedTech, Life Sci Tools, Services, CDMOs, Health Tech)
  - All companies pre-populated with ticker, name, segment
  - FactSet-friendly column headers matching DB fields
  - Blank cells for user to paste FactSet data into
  - Column groups color-coded to match the dashboard layout
  - 13 history sheets with 5 years of weekly dates (~260 weeks)

Usage:
    python -m fetcher.create_excel_template
"""

import sys
from pathlib import Path
from datetime import date, timedelta
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.company_registry import COMPANY_REGISTRY
from config.settings import DATA_DIR


# ── Column spec: (excel_header, db_field, group, description, unit_hint) ──
# Groups: ID, MARKET, FINANCIALS, MARGINS, ESTIMATES, GROWTH, PRICE
COLUMN_SPEC = [
    # ═══════════════════════════════════════════════════════════════════════
    # IDENTIFIERS — frozen reference columns, not overridable
    # ═══════════════════════════════════════════════════════════════════════
    ("Ticker",              "ticker",               "ID",         "Display ticker",                       ""),
    ("Company",             "name",                 "ID",         "Company name",                         ""),
    ("Segment",             "segment",              "ID",         "Segment key",                          ""),
    ("Sub-Segment",         "sub_segment",          "ID",         "Sub-segment (vertical only)",          ""),

    # ═══════════════════════════════════════════════════════════════════════
    # MARKET DATA — pricing, capitalization, capital structure
    # ═══════════════════════════════════════════════════════════════════════
    ("Market Cap",          "market_cap",           "MARKET",     "Market cap",                           "$M (millions)"),
    ("Enterprise Value",    "enterprise_value",     "MARKET",     "Total enterprise value",               "$M (millions)"),
    ("Stock Price",         "current_price",        "MARKET",     "Current stock price",                  "$ (raw)"),
    ("52-Week High",        "fifty_two_week_high",  "MARKET",     "52-week high price",                   "$ (raw)"),
    ("Shares Outstanding",  "shares_outstanding",   "MARKET",     "Diluted shares outstanding",           "M (millions)"),
    ("Total Debt",          "total_debt",           "MARKET",     "Total debt",                           "$M (millions)"),
    ("Total Cash",          "total_cash",           "MARKET",     "Cash & equivalents",                   "$M (millions)"),
    ("FY End Month",        "fy_end_month",         "MARKET",     "Fiscal year end month (1=Jan, 12=Dec)","1-12"),

    # ═══════════════════════════════════════════════════════════════════════
    # LTM FINANCIALS — last-twelve-months reported financials
    # ═══════════════════════════════════════════════════════════════════════
    ("LTM Revenue",         "ltm_revenue",          "FINANCIALS", "Last 12 months revenue",               "$M"),
    ("LTM Gross Profit",    "ltm_gross_profit",     "FINANCIALS", "LTM gross profit",                     "$M"),
    ("LTM EBITDA",          "ltm_ebitda",           "FINANCIALS", "LTM EBITDA",                           "$M"),
    ("LTM Operating Income","ltm_operating_income", "FINANCIALS", "LTM operating income (EBIT)",          "$M"),
    ("LTM Net Income",      "ltm_net_income",       "FINANCIALS", "LTM net income",                       "$M"),
    ("LTM Free Cash Flow",  "ltm_fcf",              "FINANCIALS", "LTM unlevered free cash flow",         "$M"),
    ("LTM R&D Expense",     "ltm_rd_expense",       "FINANCIALS", "LTM research & development",           "$M"),
    ("LTM S&M Expense",     "ltm_sm_expense",       "FINANCIALS", "LTM sales & marketing",                "$M"),
    ("LTM G&A Expense",     "ltm_ga_expense",       "FINANCIALS", "LTM general & administrative",         "$M"),
    ("LTM SBC",             "ltm_sbc",              "FINANCIALS", "LTM stock-based compensation",         "$M"),
    ("LTM CapEx",           "ltm_capex",            "FINANCIALS", "LTM capital expenditures",             "$M"),

    # ═══════════════════════════════════════════════════════════════════════
    # MARGINS — profitability & cost structure (all as decimals)
    # ═══════════════════════════════════════════════════════════════════════
    ("Gross Margin",        "gross_margin",         "MARGINS",    "Gross profit / revenue",               "decimal (0.75 = 75%)"),
    ("EBITDA Margin",       "ebitda_margin",        "MARGINS",    "EBITDA / revenue",                     "decimal (0.25 = 25%)"),
    ("Operating Margin",    "operating_margin",     "MARGINS",    "Operating income / revenue",            "decimal"),
    ("Net Margin",          "net_margin",           "MARGINS",    "Net income / revenue",                  "decimal"),
    ("FCF Margin",          "fcf_margin",           "MARGINS",    "Free cash flow / revenue",              "decimal"),
    ("R&D % Revenue",       "rd_pct_revenue",       "MARGINS",    "R&D expense / revenue",                 "decimal"),
    ("S&M % Revenue",       "sm_pct_revenue",       "MARGINS",    "Sales & marketing / revenue",           "decimal"),
    ("SBC % Revenue",       "sbc_pct_revenue",      "MARGINS",    "Stock-based comp / revenue",            "decimal"),

    # ═══════════════════════════════════════════════════════════════════════
    # CONSENSUS ESTIMATES — forward-looking analyst estimates
    # ═══════════════════════════════════════════════════════════════════════
    ("Cur FY Rev Est",      "current_fy_rev_est",   "ESTIMATES",  "Current fiscal year revenue estimate",  "$M"),
    ("Next FY Rev Est",     "next_fy_rev_est",      "ESTIMATES",  "Next fiscal year revenue estimate",     "$M"),
    ("Cur FY GP Est",       "current_fy_gp_est",    "ESTIMATES",  "Current FY gross profit estimate",      "$M"),
    ("Next FY GP Est",      "next_fy_gp_est",       "ESTIMATES",  "Next FY gross profit estimate",         "$M"),
    ("Cur FY EBITDA Est",   "current_fy_ebitda_est","ESTIMATES",  "Current FY EBITDA estimate",            "$M"),
    ("Next FY EBITDA Est",  "next_fy_ebitda_est",   "ESTIMATES",  "Next FY EBITDA estimate",               "$M"),
    ("Cur FY FCF Est",      "current_fy_fcf_est",   "ESTIMATES",  "Current FY free cash flow estimate",    "$M"),
    ("Next FY FCF Est",     "next_fy_fcf_est",      "ESTIMATES",  "Next FY free cash flow estimate",       "$M"),
    ("NTM Revenue",         "ntm_revenue",          "ESTIMATES",  "Next 12 months revenue (blended)",      "$M"),
    ("NTM Gross Profit",    "ntm_gross_profit",     "ESTIMATES",  "NTM gross profit (blended)",            "$M"),
    ("NTM EBITDA",          "ntm_ebitda",           "ESTIMATES",  "NTM EBITDA (blended)",                  "$M"),
    ("NTM FCF",             "ntm_fcf",              "ESTIMATES",  "NTM free cash flow (blended)",          "$M"),

    # ═══════════════════════════════════════════════════════════════════════
    # GROWTH — historical and forward growth metrics (all as decimals)
    # ═══════════════════════════════════════════════════════════════════════
    ("NTM Revenue Growth",  "ntm_revenue_growth",   "GROWTH",     "NTM revenue growth rate",               "decimal (0.20 = 20%)"),
    ("Cur FY Rev Growth",   "current_fy_rev_growth","GROWTH",     "Current FY revenue growth estimate",    "decimal"),
    ("Next FY Rev Growth",  "next_fy_rev_growth",   "GROWTH",     "Next FY revenue growth estimate",       "decimal"),
    ("3Y Rev CAGR",         "n3y_revenue_cagr",     "GROWTH",     "3-year revenue CAGR",                   "decimal (0.15 = 15%)"),
    ("5Y Growth Rate",      "five_year_growth_rate","GROWTH",     "5-year revenue growth rate",            "decimal"),
    ("LTM Rev Growth YoY",  "ltm_revenue_growth",  "GROWTH",     "LTM revenue growth year-over-year",     "decimal"),
    ("Net Revenue Retention","net_revenue_retention","GROWTH",     "Dollar-based net retention rate",        "decimal (1.20 = 120%)"),

    # ═══════════════════════════════════════════════════════════════════════
    # PRICE PERFORMANCE — momentum signals (all as decimals)
    # ═══════════════════════════════════════════════════════════════════════
    ("2-Week Price Chg",    "price_change_2w",      "PRICE",      "2-week price change",                   "decimal (-0.05 = -5%)"),
    ("2-Month Price Chg",   "price_change_2m",      "PRICE",      "2-month price change",                  "decimal"),
    ("6-Month Price Chg",   "price_change_6m",      "PRICE",      "6-month price change",                  "decimal"),
    ("1-Year Price Chg",    "price_change_1y",      "PRICE",      "1-year price change",                   "decimal"),
    ("YTD Price Chg",       "price_change_ytd",     "PRICE",      "Year-to-date price change",             "decimal"),
]

# Group colors for header row
GROUP_COLORS = {
    "ID":         "1F2937",  # dark gray
    "MARKET":     "1E3A5F",  # navy
    "FINANCIALS": "2D6A4F",  # forest green
    "MARGINS":    "0E7490",  # teal/cyan
    "ESTIMATES":  "7C3AED",  # purple
    "GROWTH":     "B45309",  # amber
    "PRICE":      "991B1B",  # dark red
}

GROUP_LABELS = {
    "ID":         "Identifiers",
    "MARKET":     "Market Data",
    "FINANCIALS": "LTM Financials",
    "MARGINS":    "Margins & Cost Structure",
    "ESTIMATES":  "Consensus Estimates",
    "GROWTH":     "Growth Metrics",
    "PRICE":      "Price Performance",
}

# Segment grouping
SEGMENT_SHEETS = {
    "pharma":          "Pharma",
    "consumer_health": "Consumer Health",
    "medtech":         "MedTech",
    "life_sci_tools":  "Life Sci Tools",
    "services":        "Services",
    "cdmo":            "CDMOs",
    "health_tech":     "Health Tech",
}


# ── Multiples & Time-Series History configuration ──
# Each entry: (sheet_name, db_field, color, description)
HISTORY_MULTIPLES = [
    # ── Valuation Multiples (core) ──
    ("NTM Rev x History",     "ntm_tev_rev",        "1E3A5F", "NTM TEV / Revenue"),
    ("NTM GP x History",      "ntm_tev_gp",         "2D6A4F", "NTM TEV / Gross Profit"),
    ("NTM EBITDA x History",  "ntm_tev_ebitda",     "7C3AED", "NTM TEV / EBITDA"),
    ("LTM Rev x History",     "ltm_tev_rev",        "1E3A5F", "LTM TEV / Revenue"),
    ("LTM GP x History",      "ltm_tev_gp",         "2D6A4F", "LTM TEV / Gross Profit"),
    ("LTM EBITDA x History",  "ltm_tev_ebitda",     "7C3AED", "LTM TEV / EBITDA"),

    # ── Growth-Adjusted Multiples ──
    ("GA Rev History",        "growth_adj_rev",      "B45309", "Growth-Adjusted TEV / NTM Revenue"),
    ("GA GP History",         "growth_adj_gp",       "B45309", "Growth-Adjusted TEV / NTM Gross Profit"),

    # ── Fundamentals Time-Series ──
    ("Stock Price History",   "current_price",       "991B1B", "Weekly Closing Stock Price ($)"),
    ("EV History",            "enterprise_value",    "1E3A5F", "Weekly Enterprise Value ($M)"),
    ("NTM Rev Gr History",    "ntm_revenue_growth",  "B45309", "NTM Revenue Growth Rate (decimal)"),
    ("Gross Margin History",  "gross_margin",        "2D6A4F", "Gross Margin (decimal)"),
    ("EBITDA Margin History", "ebitda_margin",       "7C3AED", "EBITDA Margin (decimal)"),
]


def _get_weekly_dates(weeks=260):
    """Generate weekly Friday dates for the trailing N weeks (default 5 years)."""
    today = date.today()
    # Find the most recent Friday (weekday 4)
    days_since_friday = (today.weekday() - 4) % 7
    last_friday = today - timedelta(days=days_since_friday)
    # If last_friday is in the future (shouldn't happen, but safety check)
    if last_friday > today:
        last_friday -= timedelta(days=7)

    dates = []
    current = last_friday
    for _ in range(weeks):
        dates.append(current)
        current -= timedelta(days=7)

    # Reverse so oldest is first (left), newest is last (right)
    dates.reverse()
    return dates


def _create_multiples_history_sheets(wb, thin_border):
    """
    Add one sheet per key multiple/metric with weekly columns for 5 years.

    Layout per sheet:
      Row 1: Sheet title + description
      Row 2: Ticker | Company | Segment | 05-Mar-21 | 12-Mar-21 | ... | 20-Feb-26
      Row 3: (ISO date hints: YYYY-MM-DD)
      Row 4+: Company data rows (blank for user to fill)
    """
    weekly_dates = _get_weekly_dates(260)  # 5 years
    # Use "DD-MMM-YY" for 5-year span to avoid ambiguity
    week_headers = [d.strftime("%d-%b-%y") for d in weekly_dates]
    # ISO date hints for machine parsing
    week_iso = [d.strftime("%Y-%m-%d") for d in weekly_dates]

    # Sort all companies by segment then name
    all_companies = sorted(COMPANY_REGISTRY, key=lambda c: (c["segment"], c["name"]))

    id_fill = PatternFill("solid", fgColor="1F2937")
    id_font = Font(color="D1D5DB", size=10)
    name_font = Font(bold=True, color="FFFFFF", size=10)

    for sheet_name, db_field, color, description in HISTORY_MULTIPLES:
        ws = wb.create_sheet(title=sheet_name)
        header_fill = PatternFill("solid", fgColor=color)
        header_font = Font(bold=True, color="FFFFFF", size=9)

        # ── Row 1: Title ──
        total_cols = 3 + len(week_headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
        title_cell = ws.cell(
            row=1, column=1,
            value=f"{description}  \u2014  Weekly History (5 Years, {len(weekly_dates)} weeks)"
        )
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = PatternFill("solid", fgColor=color)
        title_cell.alignment = Alignment(horizontal="left")

        # ── Row 2: Column headers ──
        # ID columns
        for ci, (label, width) in enumerate(
            [("Ticker", 10), ("Company", 28), ("Segment", 16)], start=1
        ):
            cell = ws.cell(row=2, column=ci, value=label)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = id_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Weekly columns — alternate fill shade by month for visual grouping
        prev_month = None
        month_toggle = False
        for wi, (header, d) in enumerate(zip(week_headers, weekly_dates)):
            ci = 4 + wi
            # Toggle shade when month changes
            if d.month != prev_month:
                month_toggle = not month_toggle
                prev_month = d.month

            cell = ws.cell(row=2, column=ci, value=header)
            cell.font = header_font
            cell.fill = header_fill if month_toggle else PatternFill("solid", fgColor="374151")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = 10

        # ── Row 3: ISO date hints ──
        for ci in range(1, 4):
            cell = ws.cell(row=3, column=ci, value="")
            cell.font = Font(italic=True, color="9CA3AF", size=8)
        for wi, iso_date in enumerate(week_iso):
            ci = 4 + wi
            cell = ws.cell(row=3, column=ci, value=iso_date)
            cell.font = Font(italic=True, color="9CA3AF", size=8)
            cell.alignment = Alignment(horizontal="center")

        # ── Data rows ──
        for row_offset, company in enumerate(all_companies):
            row = 4 + row_offset

            # Ticker
            cell = ws.cell(row=row, column=1, value=company["ticker"])
            cell.font = id_font
            cell.border = thin_border

            # Company name
            cell = ws.cell(row=row, column=2, value=company["name"])
            cell.font = name_font
            cell.border = thin_border

            # Segment
            seg_display = SEGMENT_SHEETS.get(company["segment"], company["segment"])
            cell = ws.cell(row=row, column=3, value=seg_display)
            cell.font = id_font
            cell.border = thin_border

            # Weekly columns — blank for user to fill
            for wi in range(len(week_headers)):
                ci = 4 + wi
                cell = ws.cell(row=row, column=ci)
                cell.border = thin_border

        # Freeze panes: freeze ID columns + header rows
        ws.freeze_panes = "D4"

        # Tab color — use the multiple's color
        ws.sheet_properties.tabColor = color


def create_template(output_path=None):
    """Create the FactSet override Excel template."""
    if output_path is None:
        output_path = DATA_DIR / "factset_overrides.xlsx"

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    thin_border = Border(
        bottom=Side(style="thin", color="3F3F3F"),
        right=Side(style="thin", color="3F3F3F"),
    )

    for seg_key, sheet_name in SEGMENT_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)

        companies = [c for c in COMPANY_REGISTRY if c["segment"] == seg_key]
        # Sort by name
        companies.sort(key=lambda c: c["name"])

        # ── Row 1: Group label row ──
        current_group = None
        group_start = 1
        for i, (header, db_field, group, desc, unit) in enumerate(COLUMN_SPEC):
            ci = i + 1
            if group != current_group:
                if current_group is not None:
                    # Merge the group label cells
                    if ci - 1 > group_start:
                        ws.merge_cells(
                            start_row=1, start_column=group_start,
                            end_row=1, end_column=ci - 1
                        )
                    cell = ws.cell(row=1, column=group_start)
                    cell.value = GROUP_LABELS.get(current_group, current_group)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(current_group, "333333"))
                    cell.alignment = Alignment(horizontal="center")
                current_group = group
                group_start = ci

        # Last group
        if current_group:
            last_ci = len(COLUMN_SPEC)
            if last_ci > group_start:
                ws.merge_cells(
                    start_row=1, start_column=group_start,
                    end_row=1, end_column=last_ci
                )
            cell = ws.cell(row=1, column=group_start)
            cell.value = GROUP_LABELS.get(current_group, current_group)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(current_group, "333333"))
            cell.alignment = Alignment(horizontal="center")

        # ── Row 2: Column headers ──
        for i, (header, db_field, group, desc, unit) in enumerate(COLUMN_SPEC):
            ci = i + 1
            cell = ws.cell(row=2, column=ci, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "333333"))
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # ── Row 3: Unit hints (light gray italic) ──
        for i, (header, db_field, group, desc, unit) in enumerate(COLUMN_SPEC):
            ci = i + 1
            cell = ws.cell(row=3, column=ci, value=unit if unit else "")
            cell.font = Font(italic=True, color="9CA3AF", size=8)
            cell.alignment = Alignment(horizontal="center")

        # ── Data rows: pre-fill ID columns, leave rest blank ──
        for row_offset, company in enumerate(companies):
            row = 4 + row_offset
            for i, (header, db_field, group, desc, unit) in enumerate(COLUMN_SPEC):
                ci = i + 1
                if group == "ID":
                    # Pre-fill identifier columns
                    val = company.get(db_field, "")
                    cell = ws.cell(row=row, column=ci, value=val if val else "")
                    cell.font = Font(color="D1D5DB", size=10)
                    if db_field == "name":
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                else:
                    # Leave blank for user to fill from FactSet
                    cell = ws.cell(row=row, column=ci)

                cell.border = thin_border

        # ── Column widths ──
        for i, (header, db_field, group, desc, unit) in enumerate(COLUMN_SPEC):
            ci = i + 1
            col_letter = get_column_letter(ci)
            if db_field == "name":
                ws.column_dimensions[col_letter].width = 28
            elif db_field in ("ticker", "segment", "sub_segment"):
                ws.column_dimensions[col_letter].width = 14
            elif group == "MARGINS":
                ws.column_dimensions[col_letter].width = 15
            else:
                ws.column_dimensions[col_letter].width = 18

        # Freeze panes: freeze first 4 columns (ID group) + header rows
        ws.freeze_panes = "E4"

        # Tab color
        seg_colors = {
            "pharma":          "2563EB",
            "consumer_health": "059669",
            "medtech":         "DC2626",
            "life_sci_tools":  "7C3AED",
            "services":        "F59E0B",
            "cdmo":            "EA580C",
            "health_tech":     "0891B2",
        }
        ws.sheet_properties.tabColor = seg_colors.get(seg_key, "666666")

    # ── Multiples History sheets ──
    _create_multiples_history_sheets(wb, thin_border)

    # ── Instructions sheet ──
    ws_help = wb.create_sheet(title="Instructions", index=0)
    instructions = [
        ("FactSet Override Template  \u2014  Comprehensive Edition", None),
        ("", None),
        ("HOW TO USE \u2014 OVERRIDE SHEETS (Pharma, MedTech, etc.):", None),
        ("1. Open this file in Excel with FactSet add-in enabled", None),
        ("2. Each segment sheet = one segment from the dashboard", None),
        ("3. Paste FactSet values into the colored columns", None),
        ("4. Leave cells BLANK to keep the automated pipeline value", None),
        ("5. Non-blank cells will OVERRIDE the automated pipeline values", None),
        ("6. Save the file \u2014 the dashboard reads it on every refresh", None),
        ("", None),
        ("COLUMN GROUPS (55 columns across 7 groups):", None),
        ("  Identifiers (gray)     \u2014 Pre-filled, do NOT edit (Ticker, Company, Segment, Sub-Segment)", None),
        ("  Market Data (navy)     \u2014 Market cap, EV, price, shares, debt, cash, FY end month", None),
        ("  LTM Financials (green) \u2014 Revenue, GP, EBITDA, OpInc, NetInc, FCF, R&D, S&M, G&A, SBC, CapEx", None),
        ("  Margins (teal)         \u2014 Gross, EBITDA, Operating, Net, FCF, R&D%, S&M%, SBC%", None),
        ("  Estimates (purple)     \u2014 Cur/Next FY: Rev, GP, EBITDA, FCF + NTM blended values", None),
        ("  Growth (amber)         \u2014 NTM, Cur/Next FY growth, 3Y/5Y CAGR, LTM YoY", None),
        ("  Price Perf (red)       \u2014 2-week, 2-month, 6-month, 1-year, YTD price changes", None),
        ("", None),
        ("HOW TO USE \u2014 MULTIPLES HISTORY SHEETS (13 sheets):", None),
        ("1. Thirteen history sheets track weekly data for the last 5 YEARS (~260 weeks)", None),
        ("2. Each sheet = one metric (NTM Rev x, NTM GP x, stock price, margins, etc.)", None),
        ("3. Rows = all companies, Columns = weekly Friday dates", None),
        ("4. Paste FactSet time-series data (e.g. EV/NTM Revenue by week)", None),
        ("5. Enter multiples as raw numbers (e.g. 12.5 for 12.5x)", None),
        ("6. Headers alternate color by month for easy scanning", None),
        ("7. Row 3 has ISO dates (YYYY-MM-DD) for FactSet formula references", None),
        ("8. The dashboard uses this for historical multiple trend charts", None),
        ("", None),
        ("HISTORY SHEETS:", None),
        ("  Valuation Multiples:", None),
        ("    NTM Rev x History     \u2014 NTM TEV / Revenue multiple", None),
        ("    NTM GP x History      \u2014 NTM TEV / Gross Profit multiple", None),
        ("    NTM EBITDA x History  \u2014 NTM TEV / EBITDA multiple", None),
        ("    LTM Rev x History     \u2014 LTM TEV / Revenue multiple", None),
        ("    LTM GP x History      \u2014 LTM TEV / Gross Profit multiple", None),
        ("    LTM EBITDA x History  \u2014 LTM TEV / EBITDA multiple", None),
        ("  Growth-Adjusted:", None),
        ("    GA Rev History        \u2014 Growth-adjusted TEV / NTM Rev (multiple / growth %)", None),
        ("    GA GP History         \u2014 Growth-adjusted TEV / NTM GP (multiple / growth %)", None),
        ("  Fundamentals Time-Series:", None),
        ("    Stock Price History   \u2014 Weekly closing stock price ($)", None),
        ("    EV History            \u2014 Weekly enterprise value ($M)", None),
        ("    NTM Rev Gr History    \u2014 NTM revenue growth rate (decimal)", None),
        ("    Gross Margin History  \u2014 Gross margin (decimal)", None),
        ("    EBITDA Margin History \u2014 EBITDA margin (decimal)", None),
        ("", None),
        ("UNITS:", None),
        ("  Dollar values: Enter in MILLIONS (e.g. 150000 for $150B = $150,000M)", None),
        ("    \u2192 The app multiplies by 1,000,000 to get raw dollars internally", None),
        ("    \u2192 Exception: Stock Price and 52-Week High are in raw $ (not millions)", None),
        ("  Percentages/Margins: Enter as DECIMALS (e.g. 0.25 for 25%)", None),
        ("  Multiples: Enter as raw numbers (e.g. 12.5 for 12.5x)", None),
        ("  Growth rates: Enter as DECIMALS (e.g. 0.20 for 20% growth)", None),
        ("  FY End Month: Enter as integer 1-12 (1=January, 12=December)", None),
        ("  The app handles all display formatting automatically", None),
        ("", None),
        ("TIPS:", None),
        ("  You can override just a few fields per company", None),
        ("  NTM Revenue/EBITDA are the most impactful overrides", None),
        ("  Gross Profit estimates enable more accurate GP multiples", None),
        ("  The app recalculates all multiples from the override values", None),
        ("  OpEx breakdown (R&D, S&M, G&A, SBC) enables cost structure analysis", None),
        ("  FCF data enables cash flow analysis", None),
        ("  File location: data/factset_overrides.xlsx", None),
    ]

    for row_idx, (text, _) in enumerate(instructions, start=1):
        cell = ws_help.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
        elif text.endswith(":") and text.strip():
            cell.font = Font(bold=True, size=11, color="93C5FD")
        elif text.startswith("    "):
            cell.font = Font(size=10, color="9CA3AF")
        else:
            cell.font = Font(size=10, color="D1D5DB")

    ws_help.column_dimensions["A"].width = 80

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    override_cols = len([c for c in COLUMN_SPEC if c[2] != "ID"])
    hist_sheets = len(HISTORY_MULTIPLES)
    weekly_dates = _get_weekly_dates(260)
    print(f"\u2705 Template created: {output_path}")
    print(f"   Segment sheets: {', '.join(SEGMENT_SHEETS.values())}")
    print(f"   Companies: {len(COMPANY_REGISTRY)} total")
    print(f"   Override columns: {override_cols} per segment ({len(COLUMN_SPEC)} total incl. ID)")
    print(f"   History sheets: {hist_sheets} ({len(weekly_dates)} weeks each = 5 years)")
    print(f"   Date range: {weekly_dates[0]} \u2192 {weekly_dates[-1]}")
    return output_path


if __name__ == "__main__":
    create_template()
