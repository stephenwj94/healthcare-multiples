"""
pages/drafts/comp_set_builder.py — Custom Comp Set Builder
Select 3–10 companies, view Comps Table / Regression / Trading History.
Save and load named comp sets to/from data/comp_sets.json.

Run standalone:  streamlit run pages/drafts/comp_set_builder.py
"""

from __future__ import annotations

import html as _html_lib
import io
import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent.parent))
from components.sidebar import render_sidebar
from components.formatters import (
    build_comps_df,
    compute_comps_summary,
    _cell_dollar_m,
    _cell_price_fmt,
    _cell_mult,
    _cell_52wk_html,
    _cell_rev_growth_html,
    _cell_gm_html,
    _cell_ebitda_mgn_html,
    _cell_price_change_html,
    _cell_nd,
)
from components.logos import logo_img_tag
from config.settings import DB_PATH, DATA_DIR, EXCEL_OVERRIDE_PATH
from config.color_palette import SEGMENT_SHORT as _SEG_SHORT_MAP
from fetcher.db_manager import DBManager
from fetcher.excel_override import load_overrides, apply_overrides
from utils.scatter_builder import (
    build_scatter_df, build_regression_scatter,
    SEGMENT_SHORT, SEG_COLOR_MAP, plotly_layout,
    _FONT_FAM, PLOTLY_BG, PLOTLY_GRID, PLOTLY_TEXT,
)

render_sidebar()

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&display=swap');
.block-container {
    max-width:100%!important;
    padding-left:2rem!important;
    padding-right:2rem!important;
    font-family:'DM Sans',sans-serif!important;
}
.csb-card {
    background:white;
    border:1px solid #E5E7EB;
    border-radius:12px;
    padding:16px 20px;
    box-shadow:0 1px 3px rgba(0,0,0,0.04);
    margin-bottom:12px;
}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div style="font-size:22px;font-weight:700;color:#111827;margin-bottom:2px;">Comp Set Builder</div>'
    '<div style="font-size:12px;color:#94A3B8;margin-bottom:16px;">'
    'Build and save custom peer groups — 3 to 10 companies</div>',
    unsafe_allow_html=True,
)

# ── Data ───────────────────────────────────────────────────────────────────────
db = DBManager(DB_PATH)

@st.cache_data(ttl=60)
def _load_current():
    data = db.get_all_latest_snapshots()
    ovr  = load_overrides(EXCEL_OVERRIDE_PATH)
    if ovr and data:
        data = apply_overrides(data, ovr, skip_sources={"factset"})
    return data

@st.cache_data(ttl=300)
def _load_daily(days_back=730):
    return db.get_daily_multiples(days_back=days_back)

records = _load_current()
if not records:
    st.info("No data available. Run the data fetcher to populate the database.")
    st.stop()

df_all = pd.DataFrame(records)
df_all["seg_label"] = df_all["segment"].map(SEGMENT_SHORT).fillna(df_all["segment"])

# Build ticker → display label
ticker_list = sorted(df_all["ticker"].unique().tolist())
name_map    = df_all.set_index("ticker")["name"].to_dict()
options     = [f"{t}  —  {name_map.get(t, '')}" for t in ticker_list]
ticker_from_opt = {f"{t}  —  {name_map.get(t, '')}": t for t in ticker_list}

# ── Saved comp sets ────────────────────────────────────────────────────────────
COMP_SETS_PATH = DATA_DIR / "comp_sets.json"

def _load_saved() -> dict:
    try:
        if COMP_SETS_PATH.exists():
            return json.loads(COMP_SETS_PATH.read_text())
    except Exception:
        pass
    return {}

def _save_sets(sets: dict) -> None:
    try:
        COMP_SETS_PATH.write_text(json.dumps(sets, indent=2))
    except Exception as e:
        st.error(f"Could not save comp set: {e}")

saved_sets = _load_saved()

# ── Sidebar: load / save comp sets ────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("**Saved Comp Sets**")
    if saved_sets:
        for set_name, meta in list(saved_sets.items()):
            col_a, col_b = st.columns([4, 1])
            with col_a:
                n_cos = len(meta.get("tickers", []))
                if st.button(f"📂 {set_name} ({n_cos})", key=f"load_{set_name}",
                             use_container_width=True):
                    st.session_state["csb_selected_tickers"] = meta.get("tickers", [])
                    st.rerun()
            with col_b:
                if st.button("✕", key=f"del_{set_name}", help=f"Delete '{set_name}'"):
                    del saved_sets[set_name]
                    _save_sets(saved_sets)
                    st.rerun()
    else:
        st.caption("No saved comp sets yet.")

# ── Company selector ───────────────────────────────────────────────────────────
if "csb_selected_tickers" not in st.session_state:
    st.session_state["csb_selected_tickers"] = []

saved_tickers = st.session_state["csb_selected_tickers"]
default_opts  = [o for o in options if ticker_from_opt[o] in saved_tickers]

sc1, sc2 = st.columns([7, 3])
with sc1:
    selected_opts = st.multiselect(
        "Select 3–10 companies",
        options=options,
        default=default_opts,
        max_selections=10,
        key="csb_multiselect",
    )
with sc2:
    st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
    save_name = st.text_input("Save as…", placeholder="e.g. Cloud Peers", key="csb_save_name",
                              label_visibility="collapsed")
    if st.button("Save comp set", use_container_width=True):
        if save_name.strip() and len(selected_opts) >= 3:
            tickers_to_save = [ticker_from_opt[o] for o in selected_opts]
            saved_sets[save_name.strip()] = {
                "tickers": tickers_to_save,
                "created": datetime.utcnow().isoformat(),
            }
            _save_sets(saved_sets)
            st.success(f"Saved '{save_name.strip()}'")
        elif len(selected_opts) < 3:
            st.warning("Select at least 3 companies to save.")
        else:
            st.warning("Enter a name for this comp set.")

selected_tickers = [ticker_from_opt[o] for o in selected_opts]
st.session_state["csb_selected_tickers"] = selected_tickers

if len(selected_tickers) < 3:
    st.info("Select at least 3 companies to view the comp set.")
    st.stop()

df_sel = df_all[df_all["ticker"].isin(selected_tickers)].copy()

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📋 Comps Table", "📈 Regression", "📊 Trading History"])

# ─── Comps Table shared config (matches 04_Comps_Table.py format) ─────────────

_COL1_W = 140

_CSB_COLS = [
    ("Company",           "name",          140, True,  "left"),
    ("Ticker",            "ticker",         62, True,  "left"),
    ("Segment",           "segment",        80, False, "center"),
    ("Mkt\nCap",          "mkt_cap_m",      66, False, "right"),
    ("TEV",               "tev_m",          66, False, "right"),
    ("%\n52W Hi",         "pct_52wk",       58, False, "right"),
    ("NTM\nEV/Rev",       "ev_rev",         62, False, "right"),
    ("LTM\nEV/Rev",       "ltm_rev_x",      62, False, "right"),
    ("NTM\nEV/GP",        "ev_gp",          62, False, "right"),
    ("LTM\nEV/GP",        "ltm_gp_x",       62, False, "right"),
    ("NTM\nEV/EBITDA",    "ev_ebitda",      66, False, "right"),
    ("LTM\nEV/EBITDA",    "ltm_ebitda_x",   66, False, "right"),
    ("NTM Rev\nGr%",      "rev_gr",          62, False, "right"),
    ("3Y\nCAGR",          "cagr_3y",         58, False, "right"),
    ("Gross\nMgn",        "gm",              58, False, "right"),
    ("NTM EBITDA\nMgn",   "ebitda_mgn",      68, False, "right"),
    ("1M\nChg",           "chg_2w",          58, False, "right"),
    ("1Y\nChg",           "chg_2m",          58, False, "right"),
]

_CSB_GROUPS = [
    ("",                   2, "#1E293B",     "transparent", "transparent"),
    ("Segment",            1, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Market Data",        3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Revenue",            2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Gross Profit",       2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("EBITDA",             2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Growth & Margins",   4, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Price Performance",  2, "#1E293B",     "#FFFFFF",     "transparent"),
]

_CSB_GROUP_STARTS = {2, 3, 6, 8, 10, 12, 16}
_LTM_KEYS = {"ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}
_MULT_KEYS = {"ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}

_CSB_CFG = {
    "cols":         _CSB_COLS,
    "groups":       _CSB_GROUPS,
    "group_starts": _CSB_GROUP_STARTS,
}

SEGMENT_PILL_COLORS = {
    "pharma":          ("#DBEAFE", "#1D4ED8"),
    "consumer_health": ("#D1FAE5", "#065F46"),
    "medtech":         ("#FEE2E2", "#991B1B"),
    "life_sci_tools":  ("#EDE9FE", "#5B21B6"),
    "services":        ("#FEF3C7", "#92400E"),
    "cdmo":            ("#CFFAFE", "#155E75"),
    "health_tech":     ("#FCE7F3", "#9D174D"),
}

_BG_BASE  = "#FFFFFF"
_BG_ALT   = "#FAFBFD"
_BG_HOVER = "#F0F4FF"
_BG_SUM   = "#E2E8F0"

_SHORT_NAMES = {
    "Zoom Video Communications, Inc.": "Zoom",
    "Zoom Video Communications":       "Zoom",
    "Salesforce, Inc.":                "Salesforce",
    "ServiceNow, Inc.":                "ServiceNow",
}

def _csb_cell_segment_pill(segment_key):
    short = _SEG_SHORT_MAP.get(str(segment_key), str(segment_key))
    bg, fg = SEGMENT_PILL_COLORS.get(str(segment_key), ("#374151", "#D1D5DB"))
    safe = _html_lib.escape(short)
    return (
        f'<span style="background:{bg};color:{fg};padding:2px 6px;'
        f'border-radius:4px;font-size:10px;font-weight:600;white-space:nowrap;">'
        f'{safe}</span>'
    )

def _csb_render_cell(key, val, row_dict=None):
    if key == "name":
        full_name = str(val or "")
        disp_name = _SHORT_NAMES.get(full_name, full_name)
        safe_disp = _html_lib.escape(disp_name)
        return (
            f'<span style="font-weight:500;color:#111827;white-space:nowrap;'
            f'display:inline-block;max-width:132px;overflow:hidden;'
            f'text-overflow:ellipsis">{safe_disp}</span>'
        )
    if key == "ticker":
        safe = _html_lib.escape(str(val or ""))
        logo = logo_img_tag(str(val or ""), size=14)
        logo_span = (
            f'<span style="display:inline-flex;align-items:center;'
            f'width:14px;height:14px;margin-right:3px;flex-shrink:0;">{logo}</span>'
            if logo else ""
        )
        return (
            f'<span style="display:inline-flex;align-items:center;font-weight:600;'
            f'font-size:10px;color:#111827;">'
            f'{logo_span}<span>{safe}</span></span>'
        )
    if key == "segment":
        return _csb_cell_segment_pill(val)
    if key in ("mkt_cap_m", "tev_m"):
        return _cell_dollar_m(val)
    if key == "pct_52wk":
        return _cell_52wk_html(val)
    if key in ("ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"):
        return _cell_mult(val)
    if key in ("rev_gr", "cagr_3y"):
        return _cell_rev_growth_html(val)
    if key == "gm":
        return _cell_gm_html(val)
    if key == "ebitda_mgn":
        return _cell_ebitda_mgn_html(val)
    if key in ("chg_2w", "chg_2m"):
        return _cell_price_change_html(val)
    if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
        return _cell_nd()
    return _html_lib.escape(str(val))

def _csb_sort_val(key, val):
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    if key in _MULT_KEYS and isinstance(val, (int, float)) and val <= 0:
        return ""
    if isinstance(val, (int, float)):
        return str(float(val))
    return _html_lib.escape(str(val).lower())

def _csb_build_table_html(df, mean_d, median_d, cfg):
    cols         = cfg["cols"]
    groups       = cfg["groups"]
    group_starts = cfg["group_starts"]

    css = """<style>
.csb-outer { overflow-x:auto; border:1px solid #D1D5DB; border-radius:4px; margin-bottom:4px;
    font-family:'DM Sans',-apple-system,BlinkMacSystemFont,'Inter','Segoe UI',sans-serif; background:%(base)s; }
.csb-tbl { border-collapse:separate; border-spacing:0; width:100%%; font-size:11px;
    font-variant-numeric:tabular-nums; background:%(base)s; }
.csb-thead { position:sticky; top:0; z-index:22; background:#1E293B; }
.csb-gr { padding:4px 6px; font-size:9px; font-weight:700; letter-spacing:0.07em;
    text-transform:uppercase; text-align:center; white-space:nowrap; background:#1E293B!important; color:#FFFFFF; }
.csb-ch { padding:6px 4px; font-size:9.5px; font-weight:700; color:#FFFFFF; text-transform:uppercase;
    letter-spacing:0.03em; background:#1E293B!important; text-align:center; border-bottom:none;
    white-space:normal; line-height:1.3; vertical-align:bottom; }
.csb-gs { border-left:1px solid #CBD5E1; }
.csb-thead .csb-gs { border-left:none; box-shadow:inset 1px 0 0 #334155; }
.csb-td { padding:2px 5px; text-align:right; border-bottom:1px solid #F3F4F6;
    background:%(base)s; vertical-align:middle; white-space:nowrap; color:#111827; line-height:1.2; }
.csb-td.lft { text-align:left; }
.csb-td.ctr { text-align:center; }
.csb-tr:hover .csb-td { background-color:%(hover)s!important; }
.csb-alt .csb-td { background:%(alt)s; }
.csb-alt:hover .csb-td { background-color:%(hover)s!important; }
.csb-sum .csb-td { background:%(sum)s!important; font-weight:600; font-size:11px; padding:2px 5px;
    border-bottom:1px solid #E2E8F0; color:#111827; }
.csb-sum .csb-td span { color:#111827!important; }
.csb-mean .csb-td { border-top:1px solid #94A3B8!important; }
.csb-med .csb-td { border-bottom:2px solid #94A3B8!important; }
.csb-ltm.csb-td { color:#111827; }
.csb-s0 { position:sticky; left:0; z-index:6; }
.csb-s1 { position:sticky; left:%(col1w)spx; z-index:6; border-right:1px solid #E5E7EB; }
.csb-alt .csb-td.csb-s0, .csb-alt .csb-td.csb-s1 { background:%(alt)s; }
.csb-sum .csb-td.csb-s0, .csb-sum .csb-td.csb-s1 { background:%(sum)s!important; }
.csb-tr:hover .csb-td.csb-s0, .csb-tr:hover .csb-td.csb-s1,
.csb-alt:hover .csb-td.csb-s0, .csb-alt:hover .csb-td.csb-s1 { background-color:%(hover)s!important; }
.csb-ch[data-col] { cursor:pointer; user-select:none; }
.csb-ch[data-col]::after { content:"\\2195"; color:#D1D5DB; display:block; font-size:7px; line-height:1; margin-top:1px; text-align:center; }
</style>""" % {"base": _BG_BASE, "alt": _BG_ALT, "hover": _BG_HOVER, "sum": _BG_SUM, "col1w": _COL1_W}

    # thead
    gr_cells = []
    for gname, colspan, bg, fg, _ in groups:
        label = _html_lib.escape(gname)
        gr_cells.append(f'<th class="csb-gr" colspan="{colspan}">{label}</th>')
    col_cells = []
    for i, (label, key, width, sticky, align) in enumerate(cols):
        s_cls = (" csb-s0" if sticky and i == 0 else " csb-s1" if sticky and i == 1 else "")
        l_sty = ("left:0px;" if i == 0 and sticky else f"left:{_COL1_W}px;" if i == 1 and sticky else "")
        gs = " csb-gs" if i in group_starts else ""
        ltm = " csb-ltm" if key in _LTM_KEYS else ""
        disp = '<br>'.join(_html_lib.escape(p) for p in label.split('\n'))
        col_cells.append(
            f'<th class="csb-ch{s_cls}{gs}{ltm}" data-col="{_html_lib.escape(key)}" '
            f'style="min-width:{width}px;{l_sty}">{disp}</th>'
        )
    thead = f'<thead class="csb-thead"><tr>{"".join(gr_cells)}</tr><tr>{"".join(col_cells)}</tr></thead>'

    # summary row helper
    def _sum_row(vals_d, label, extra_cls=""):
        cells = []
        for i, (_, key, width, sticky, align) in enumerate(cols):
            s_cls = (" csb-s0" if sticky and i == 0 else " csb-s1" if sticky and i == 1 else "")
            l_sty = ("left:0px;" if i == 0 and sticky else f"left:{_COL1_W}px;" if i == 1 and sticky else "")
            a_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
            gs = " csb-gs" if i in group_starts else ""
            ltm = " csb-ltm" if key in _LTM_KEYS else ""
            if key == "name":
                content = f'<span style="color:#94A3B8;font-style:italic;font-size:11px;font-weight:400">{_html_lib.escape(label)}</span>'
            elif key in ("ticker", "mkt_cap_m", "tev_m", "segment"):
                content = ""
            else:
                content = _csb_render_cell(key, vals_d.get(key))
            cells.append(f'<td class="csb-td{s_cls}{a_cls}{gs}{ltm}" style="{l_sty}">{content}</td>')
        return f'<tr class="csb-tr csb-sum{extra_cls}">{"".join(cells)}</tr>'

    # data rows
    data_rows = []
    for idx, row_dict in enumerate(df.to_dict("records")):
        alt = " csb-alt" if idx % 2 == 1 else ""
        cells = []
        for i, (_, key, width, sticky, align) in enumerate(cols):
            s_cls = (" csb-s0" if sticky and i == 0 else " csb-s1" if sticky and i == 1 else "")
            l_sty = ("left:0px;" if i == 0 and sticky else f"left:{_COL1_W}px;" if i == 1 and sticky else "")
            a_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
            gs = " csb-gs" if i in group_starts else ""
            ltm = " csb-ltm" if key in _LTM_KEYS else ""
            val = row_dict.get(key)
            content = _csb_render_cell(key, val, row_dict)
            sv = _csb_sort_val(key, val)
            cells.append(f'<td class="csb-td{s_cls}{a_cls}{gs}{ltm}" data-sort="{sv}" style="{l_sty}">{content}</td>')
        data_rows.append(f'<tr class="csb-tr{alt}">{"".join(cells)}</tr>')

    return (
        css
        + '<div class="csb-outer"><table class="csb-tbl">'
        + thead + "<tbody>"
        + _sum_row(mean_d, "Mean", " csb-mean")
        + _sum_row(median_d, "Median", " csb-med")
        + "".join(data_rows)
        + "</tbody></table></div>"
    )


def _csb_build_excel(df, mean_d, median_d):
    """Build an Excel file for the custom comp set."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    _EXCEL_COLUMNS = [
        ("Company",          "name",          28, None),
        ("Ticker",           "ticker",        10, None),
        ("Segment",          "segment",       16, None),
        ("Mkt Cap ($mm)",    "mkt_cap_m",     14, '#,##0;(#,##0)'),
        ("TEV ($mm)",        "tev_m",         14, '#,##0;(#,##0)'),
        ("% 52W Hi",         "pct_52wk",      10, '0%;(0%)'),
        ("NTM EV/Rev",       "ev_rev",        11, '0.0"x";(0.0"x")'),
        ("LTM EV/Rev",       "ltm_rev_x",     11, '0.0"x";(0.0"x")'),
        ("NTM EV/GP",        "ev_gp",         11, '0.0"x";(0.0"x")'),
        ("LTM EV/GP",        "ltm_gp_x",      11, '0.0"x";(0.0"x")'),
        ("NTM EV/EBITDA",    "ev_ebitda",     13, '0.0"x";(0.0"x")'),
        ("LTM EV/EBITDA",    "ltm_ebitda_x",  13, '0.0"x";(0.0"x")'),
        ("NTM Rev Gr%",      "rev_gr",        12, '0.0%;(0.0%)'),
        ("3Y CAGR",          "cagr_3y",       10, '0.0%;(0.0%)'),
        ("Gross Mgn",        "gm",            10, '0%;(0%)'),
        ("NTM EBITDA Mgn",   "ebitda_mgn",    14, '0%;(0%)'),
        ("1M Chg",           "chg_2w",        10, '0.0%;(0.0%)'),
        ("1Y Chg",           "chg_2m",        10, '0.0%;(0.0%)'),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comp Set"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sum_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    sum_font = Font(name="Calibri", bold=True, size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font_black = Font(name="Calibri", size=10, color="111827")
    data_font_blue = Font(name="Calibri", size=10, color="0000FF")
    data_align_r = Alignment(horizontal="right", vertical="center")
    data_align_l = Alignment(horizontal="left", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))
    thick_border = Border(bottom=Side(style="medium", color="334155"))

    ws.cell(row=1, column=1, value="Custom Comp Set")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Data as of {date.today().strftime('%B %d, %Y')}. All financials in $mm.")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="6B7280")

    hdr_row = 4
    for col_idx, (header, key, width, fmt) in enumerate(_EXCEL_COLUMNS, 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thick_border
        ws.column_dimensions[cell.column_letter].width = width

    def _write_sum(row_num, label, summary_dict):
        for col_idx, (header, key, width, fmt) in enumerate(_EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = sum_fill
            cell.font = sum_font
            cell.border = thin_border
            cell.alignment = data_align_r
            if col_idx == 1:
                cell.value = label
                cell.alignment = data_align_l
            elif col_idx <= 3:
                continue
            else:
                val = summary_dict.get(key)
                if val is not None and not (isinstance(val, float) and (val != val)):
                    cell.value = val
                    if fmt:
                        cell.number_format = fmt

    _write_sum(hdr_row + 1, "Mean", mean_d)
    _write_sum(hdr_row + 2, "Median", median_d)

    data_start = hdr_row + 3
    for i, (_, row) in enumerate(df.iterrows()):
        row_num = data_start + i
        is_alt = i % 2 == 1
        for col_idx, (header, key, width, fmt) in enumerate(_EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            val = row.get(key)
            if key == "segment" and val:
                val = _SEG_SHORT_MAP.get(str(val), str(val))
            if val is None or (isinstance(val, float) and val != val):
                cell.font = data_font_black
                cell.value = None
                continue
            cell.value = val
            if col_idx <= 3:
                cell.font = data_font_black
                cell.alignment = data_align_l
            else:
                cell.font = data_font_blue
                cell.alignment = data_align_r
                if fmt:
                    cell.number_format = fmt

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Tab 1: Comps Table ───────────────────────────────────────────────────────
with tab1:
    # Build comps DataFrame using the same formatter as the main Comps Table
    sel_records = df_sel.to_dict("records")
    df_comps = build_comps_df(sel_records)

    if df_comps.empty:
        st.warning("No data to display for selected companies.")
    else:
        mean_d, median_d = compute_comps_summary(df_comps)

        # Download Comp Table button
        xlsx_data = _csb_build_excel(df_comps, mean_d, median_d)
        st.download_button(
            label="Download Comp Table",
            data=xlsx_data,
            file_name="custom_comp_set.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        table_html = _csb_build_table_html(df_comps, mean_d, median_d, _CSB_CFG)
        st.markdown(table_html, unsafe_allow_html=True)

# ─── Tab 2: Regression ────────────────────────────────────────────────────────
with tab2:
    rc1, rc2 = st.columns(2)
    with rc1:
        x_opt = st.selectbox("X axis", ["NTM EV/Revenue", "NTM EV/EBITDA"], key="csb_x")
    with rc2:
        y_opt = st.selectbox("Y axis", ["NTM Revenue Growth %", "NTM EBITDA Margin %"],
                             key="csb_y")

    _x_map = {"NTM EV/Revenue": "NTM Rev x", "NTM EV/EBITDA": "NTM EBITDA x"}
    _y_map = {"NTM Revenue Growth %": "NTM Rev Growth",
              "NTM EBITDA Margin %": "EBITDA Margin"}
    _y_sfx = {"NTM Revenue Growth %": "%", "NTM EBITDA Margin %": "%"}

    df_scatter = build_scatter_df(df_sel.to_dict("records"))
    fig = build_regression_scatter(
        df_scatter, _x_map[x_opt], _y_map[y_opt],
        x_label=x_opt + " (x)",
        y_label=y_opt,
        height=520,
        x_suffix="x",
        y_suffix=_y_sfx[y_opt],
    )
    if fig:
        st.markdown('<div class="csb-card">', unsafe_allow_html=True)
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.caption("Not enough data for regression scatter.")

# ─── Tab 3: Trading History ───────────────────────────────────────────────────
with tab3:
    th1, th2 = st.columns([3, 7])
    with th1:
        hist_metric = st.selectbox(
            "Metric",
            ["NTM TEV/Revenue", "NTM TEV/EBITDA"],
            key="csb_hist_metric",
        )
    _HIST_COL = {"NTM TEV/Revenue": "ntm_tev_rev", "NTM TEV/EBITDA": "ntm_tev_ebitda"}
    hist_col = _HIST_COL[hist_metric]

    daily_records = _load_daily(730)
    if not daily_records:
        st.caption("No daily multiples history available.")
    else:
        df_dm = pd.DataFrame(daily_records)
        df_dm["_dt"] = pd.to_datetime(df_dm["date"], errors="coerce")
        df_dm = df_dm[df_dm["ticker"].isin(selected_tickers)].dropna(subset=["_dt", hist_col])

        if df_dm.empty:
            st.caption("No historical data for selected companies.")
        else:
            fig_ts = go.Figure()
            layout = plotly_layout(height=520)
            layout["hovermode"] = "x unified"
            layout["yaxis"]["ticksuffix"] = "x"
            layout["legend"] = dict(
                orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                font=dict(size=10, family=_FONT_FAM),
                bgcolor="rgba(0,0,0,0)",
            )
            layout["xaxis"]["showgrid"] = False
            layout["margin"].update({"r": 60, "t": 55})

            _COLORS = [
                "#3B82F6", "#16A34A", "#F59E0B", "#EF4444", "#8B5CF6",
                "#EC4899", "#06B6D4", "#F97316", "#84CC16", "#6366F1",
            ]

            for i, ticker in enumerate(selected_tickers):
                sub = df_dm[df_dm["ticker"] == ticker].sort_values("_dt")
                if sub.empty:
                    continue
                # 7-day rolling smoothing
                sub = sub.set_index("_dt")[hist_col].rolling(window=7, min_periods=2).mean().reset_index()
                sub.columns = ["_dt", hist_col]
                sub = sub.dropna()

                color = _COLORS[i % len(_COLORS)]
                d_strs = [pd.Timestamp(d).strftime("%Y-%m-%d") for d in sub["_dt"]]
                fig_ts.add_trace(go.Scatter(
                    x=d_strs,
                    y=sub[hist_col].tolist(),
                    mode="lines",
                    name=ticker,
                    line=dict(color=color, width=2),
                    hovertemplate=f"{ticker}: %{{y:.1f}}x<extra></extra>",
                ))

            fig_ts.update_layout(**layout)

            st.markdown('<div class="csb-card">', unsafe_allow_html=True)
            st.plotly_chart(fig_ts, use_container_width=True, config={"displayModeBar": False})
            st.markdown('</div>', unsafe_allow_html=True)
