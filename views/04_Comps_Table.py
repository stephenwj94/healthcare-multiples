"""
Consolidated Comps Table -- all 7 healthcare segments in one page with toggle pills.
"""

import html as _html_lib
import io
import streamlit as st
import streamlit.components.v1 as _st_comps
import numpy as np
import pandas as pd
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))
from components.sidebar import render_sidebar
from components.formatters import (
    build_comps_df,
    compute_comps_summary,
    _cell_dollar_m,
    _cell_price_fmt,
    _cell_mult,
    _cell_52wk_html,
    _cell_rev_growth_html,
    _cell_gm_html,
    _cell_ebitda_mgn_html,
    _cell_price_change_html,
    _cell_nd,
)
from config.settings import DB_PATH, SEGMENT_DISPLAY, EXCEL_OVERRIDE_PATH
from config.color_palette import SEGMENT_COLORS, SEGMENT_SHORT, LIGHT_BADGE_STYLES
from fetcher.db_manager import DBManager
from fetcher.excel_override import load_overrides, apply_overrides
from components.logos import logo_img_tag

render_sidebar()

# -- Page header ---------------------------------------------------------------

st.title("Comps Table")
st.markdown(
    '<p style="color:#94A3B8;font-size:13px;margin:-8px 0 16px 0;">'
    "All healthcare segments in one view. Click column headers to sort. "
    "All financials in millions ($mm).</p>",
    unsafe_allow_html=True,
)

# -- Data loading --------------------------------------------------------------

db = DBManager(DB_PATH)

# -- Segment toggles (checkboxes) ---------------------------------------------

SEGMENTS = list(SEGMENT_DISPLAY.keys())
SEGMENT_LABELS = {k: SEGMENT_SHORT.get(k, v) for k, v in SEGMENT_DISPLAY.items()}

# Initialize all segments as selected
if "comps_segments" not in st.session_state:
    st.session_state["comps_segments"] = set(SEGMENTS)

selected = st.session_state["comps_segments"]

# Build toggle checkboxes in a single row
chk_cols = st.columns(len(SEGMENTS) + 1)  # extra col for "All/Clear" toggle

with chk_cols[0]:
    all_selected = len(selected) == len(SEGMENTS)
    if st.button(
        "All" if not all_selected else "Clear",
        key="comps_toggle_all",
        type="secondary",
    ):
        if all_selected:
            st.session_state["comps_segments"] = set()
        else:
            st.session_state["comps_segments"] = set(SEGMENTS)
        st.rerun()

for i, seg_key in enumerate(SEGMENTS):
    label = SEGMENT_LABELS[seg_key]
    is_on = seg_key in selected
    with chk_cols[i + 1]:
        new_val = st.checkbox(label, value=is_on, key=f"comps_chk_{seg_key}")
        if new_val and seg_key not in st.session_state["comps_segments"]:
            st.session_state["comps_segments"].add(seg_key)
            st.rerun()
        elif not new_val and seg_key in st.session_state["comps_segments"]:
            st.session_state["comps_segments"].discard(seg_key)
            st.rerun()

selected = st.session_state["comps_segments"]

if not selected:
    st.info("Select at least one segment to display the comps table.")
    st.stop()

# -- Load & filter data --------------------------------------------------------

all_data = db.get_all_latest_snapshots()

# Apply FactSet Excel overrides
@st.cache_data(ttl=60)
def _cached_load_overrides():
    return load_overrides(EXCEL_OVERRIDE_PATH)

overrides = _cached_load_overrides()
if overrides:
    all_data = apply_overrides(all_data, overrides, skip_sources={"factset"})

# Filter to selected segments
filtered_data = [d for d in all_data if d.get("segment") in selected]

if not filtered_data:
    st.warning("No data available for the selected segments. Run the data fetcher to populate the database.")
    st.stop()

# -- Build DataFrame -----------------------------------------------------------

df = build_comps_df(filtered_data)
if df.empty:
    st.warning("No data to display.")
    st.stop()

mean_d, median_d = compute_comps_summary(df)

# -- Column definitions --------------------------------------------------------
# Reduced widths to minimize horizontal scrolling.
# Key multiples visible without scrolling; price perf at far right.

_COL1_W = 140

_COLS = [
    # Info group (sticky)
    ("Company",           "name",          140, True,  "left"),
    ("Ticker",            "ticker",         62, True,  "left"),
    # Segment pill
    ("Segment",           "segment",        80, False, "center"),
    # Market Data
    ("Mkt\nCap",          "mkt_cap_m",      66, False, "right"),
    ("TEV",               "tev_m",          66, False, "right"),
    ("%\n52W Hi",         "pct_52wk",       58, False, "right"),
    # Revenue multiples
    ("NTM\nEV/Rev",       "ev_rev",         62, False, "right"),
    ("LTM\nEV/Rev",       "ltm_rev_x",      62, False, "right"),
    # Gross Profit multiples
    ("NTM\nEV/GP",        "ev_gp",          62, False, "right"),
    ("LTM\nEV/GP",        "ltm_gp_x",       62, False, "right"),
    # EBITDA multiples
    ("NTM\nEV/EBITDA",    "ev_ebitda",      66, False, "right"),
    ("LTM\nEV/EBITDA",    "ltm_ebitda_x",   66, False, "right"),
    # Growth & Margins
    ("NTM Rev\nGr%",      "rev_gr",          62, False, "right"),
    ("3Y\nCAGR",          "cagr_3y",         58, False, "right"),
    ("Gross\nMgn",        "gm",              58, False, "right"),
    ("NTM EBITDA\nMgn",   "ebitda_mgn",      68, False, "right"),
    # Price Performance (relabeled)
    ("1M\nChg",           "chg_2w",          58, False, "right"),
    ("1Y\nChg",           "chg_2m",          58, False, "right"),
]

_GROUPS = [
    ("",                   2, "#1E293B",     "transparent", "transparent"),
    ("Segment",            1, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Market Data",        3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Revenue",            2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Gross Profit",       2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("EBITDA",             2, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Growth & Margins",   4, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Price Performance",  2, "#1E293B",     "#FFFFFF",     "transparent"),
]

# Col indices where group-divider left-border is applied
_GROUP_STARTS = {2, 3, 6, 8, 10, 12, 16}

_CFG = {
    "cols":         _COLS,
    "groups":       _GROUPS,
    "group_starts": _GROUP_STARTS,
    "tev_idx":      4,
}

# -- LTM column keys -- receive lighter text styling --------------------------
_LTM_KEYS = {"ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}

# -- Multiple column keys -- non-positive values sort last --------------------
_MULT_KEYS = {"ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}

# -- Short display names (full DB name -> condensed label) ---------------------
_SHORT_NAMES = {
    "Zoom Video Communications, Inc.": "Zoom",
    "Zoom Video Communications":       "Zoom",
    "Salesforce, Inc.":                "Salesforce",
    "ServiceNow, Inc.":                "ServiceNow",
}

# -- Pending / rumored M&A annotations ----------------------------------------
_MA_NOTES = {
    "OS": "Pending take-private -- Hg Capital ($6.4B all-cash; expected H1 2026)",
}

# -- Segment pill badge --------------------------------------------------------

SEGMENT_PILL_COLORS = {
    "pharma":          ("#DBEAFE", "#1D4ED8"),
    "consumer_health": ("#D1FAE5", "#065F46"),
    "medtech":         ("#FEE2E2", "#991B1B"),
    "life_sci_tools":  ("#EDE9FE", "#5B21B6"),
    "services":        ("#FEF3C7", "#92400E"),
    "cdmo":            ("#CFFAFE", "#155E75"),
    "health_tech":     ("#FCE7F3", "#9D174D"),
}


def _cell_segment_pill(segment_key):
    """Render a colored pill badge for the segment."""
    short = SEGMENT_SHORT.get(str(segment_key), str(segment_key))
    bg, fg = SEGMENT_PILL_COLORS.get(str(segment_key), ("#374151", "#D1D5DB"))
    safe = _html_lib.escape(short)
    return (
        f'<span style="background:{bg};color:{fg};padding:2px 6px;'
        f'border-radius:4px;font-size:10px;font-weight:600;white-space:nowrap;">'
        f'{safe}</span>'
    )


# -- Background colors ---------------------------------------------------------
_BG_BASE  = "#FFFFFF"
_BG_ALT   = "#FAFBFD"
_BG_HOVER = "#F0F4FF"
_BG_SUM   = "#E2E8F0"

# -- CSS -----------------------------------------------------------------------
_CT_CSS = """<style>
.ct-outer {
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    border: 1px solid #D1D5DB;
    border-radius: 4px;
    margin-bottom: 4px;
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, 'Inter', 'Segoe UI', sans-serif;
    background: %(base)s;
}
.ct-tbl {
    border-collapse: separate;
    border-spacing: 0;
    width: 100%%;
    font-size: 11px;
    font-variant-numeric: tabular-nums;
    background: %(base)s;
}
.ct-thead { position: sticky; top: 0; z-index: 22; background: #1E293B; }
.ct-gr {
    padding: 4px 6px;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 0.07em;
    text-transform: uppercase;
    text-align: center;
    white-space: nowrap;
}
.ct-ch {
    padding: 6px 4px 6px;
    font-size: 9.5px;
    font-weight: 700;
    color: #FFFFFF;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    background: #1E293B;
    text-align: center;
    border-bottom: none;
    white-space: normal;
    line-height: 1.3;
    vertical-align: bottom;
}
.ct-ch.lft { text-align: left; }
.ct-ch.ctr { text-align: center; }
.ct-ch.ct-s0, .ct-ch.ct-s1 { z-index: 26; background: #1E293B; }
.ct-gs { border-left: 1px solid #CBD5E1; }
.ct-thead .ct-gs { border-left: none; box-shadow: inset 1px 0 0 #334155; }
.ct-gr, .ct-ch { background: #1E293B !important; }
.ct-thead tr { background: #1E293B !important; }
.ct-thead th { background: #1E293B !important; }
.ct-td {
    padding: 2px 5px;
    text-align: right;
    border-bottom: 1px solid #F3F4F6;
    background: %(base)s;
    vertical-align: middle;
    white-space: nowrap;
    color: #111827;
    line-height: 1.2;
}
.ct-td.lft { text-align: left; }
.ct-td.ctr { text-align: center; }
.ct-tr:hover .ct-td      { background-color: %(hover)s !important; }
.ct-alt .ct-td            { background: %(alt)s; }
.ct-alt:hover .ct-td      { background-color: %(hover)s !important; }
.ct-sum .ct-td {
    background: %(sum)s !important;
    font-weight: 600;
    font-size: 11px;
    padding: 2px 5px;
    border-bottom: 1px solid #E2E8F0;
    color: #111827;
}
.ct-sum .ct-td span { color: #111827 !important; }
.ct-mean .ct-td { border-top: 1px solid #94A3B8 !important; }
.ct-med .ct-td  { border-bottom: 2px solid #94A3B8 !important; }
.ct-ltm.ct-td { color: #111827; }
.ct-ltm.ct-ch { color: #FFFFFF; }
.ct-s0 { position: sticky; left: 0;        z-index: 6; }
.ct-s1 { position: sticky; left: %(col1w)spx;    z-index: 6; border-right: 1px solid #E5E7EB; }
.ct-alt .ct-td.ct-s0,
.ct-alt .ct-td.ct-s1       { background: %(alt)s; }
.ct-sum .ct-td.ct-s0,
.ct-sum .ct-td.ct-s1       { background: %(sum)s !important; }
.ct-tr:hover .ct-td.ct-s0,
.ct-tr:hover .ct-td.ct-s1,
.ct-alt:hover .ct-td.ct-s0,
.ct-alt:hover .ct-td.ct-s1 { background-color: %(hover)s !important; }
.ct-tkr { color:#111827; font-weight:600; font-size:10px; text-decoration:none; }
.ct-tkr:hover { text-decoration:underline; }
.ct-ch[data-col] { cursor: pointer; user-select: none; }
.ct-ch[data-col]:hover { color: #FFFFFF; }
.ct-ch[data-col]::after {
    content: "\\2195";
    color: #D1D5DB;
    display: block;
    font-size: 7px;
    line-height: 1;
    margin-top: 1px;
    text-align: center;
}
.ct-ch.ct-sort-asc  { color: #3B82F6 !important; }
.ct-ch.ct-sort-desc { color: #3B82F6 !important; }
.ct-ch.ct-sort-asc::after  { content: "\\25B2"; color: #3B82F6; display: block; font-size: 7px; line-height: 1; margin-top: 1px; text-align: center; }
.ct-ch.ct-sort-desc::after { content: "\\25BC"; color: #3B82F6; display: block; font-size: 7px; line-height: 1; margin-top: 1px; text-align: center; }
/* Company name link styling */
a.ct-name-link { text-decoration: none; color: inherit; }
a.ct-name-link:hover { text-decoration: underline; }
</style>""" % {"base": _BG_BASE, "alt": _BG_ALT, "hover": _BG_HOVER, "sum": _BG_SUM, "col1w": _COL1_W}


# -- Sortable-column JS -------------------------------------------------------
_STICKY_JS = """<script>
(function(){
  try {
    var win = window.parent;
    var doc = win.document;

    if (!win.__ctSortState) win.__ctSortState = {};
    var _sortState = win.__ctSortState;

    function _cmp(aSort, bSort, dir) {
      if (aSort === '' && bSort === '') return 0;
      if (aSort === '') return 1;
      if (bSort === '') return -1;
      var an = parseFloat(aSort), bn = parseFloat(bSort);
      var v  = (isNaN(an) || isNaN(bn)) ? aSort.localeCompare(bSort) : an - bn;
      return dir === 'asc' ? v : -v;
    }

    function _sortTable(table, ths, colIdx, dir) {
      var tbody = table.querySelector('tbody');
      if (!tbody) return;
      var all  = Array.from(tbody.querySelectorAll('tr'));
      var summ = all.filter(function(r){ return r.classList.contains('ct-sum'); });
      var data = all.filter(function(r){ return !r.classList.contains('ct-sum'); });
      data.sort(function(a, b) {
        var tds = function(r){ return r.querySelectorAll('td'); };
        var ac  = tds(a)[colIdx], bc = tds(b)[colIdx];
        return _cmp(
          ac ? (ac.dataset.sort || '') : '',
          bc ? (bc.dataset.sort || '') : '',
          dir
        );
      });
      summ.concat(data).forEach(function(r){ tbody.appendChild(r); });
      data.forEach(function(r, i){
        r.classList.toggle('ct-alt', i % 2 === 1);
      });
      ths.forEach(function(th, i){
        th.classList.toggle('ct-sort-asc',  i === colIdx && dir === 'asc');
        th.classList.toggle('ct-sort-desc', i === colIdx && dir === 'desc');
      });
    }

    function _initTableSort(table) {
      if (table.__ctSortBound) return;
      table.__ctSortBound = true;
      var id = 'ct_' + Math.random().toString(36).substr(2,9);
      table.__ctId = id;
      _sortState[id] = {colIdx: -1, dir: 'desc'};
      var thead = table.querySelector('thead.ct-thead');
      if (!thead) return;
      var hrows = thead.querySelectorAll('tr');
      if (hrows.length < 2) return;
      var ths = Array.from(hrows[1].querySelectorAll('th[data-col]'));
      ths.forEach(function(th, idx){
        th.addEventListener('click', function(){
          var st  = _sortState[id];
          var dir = (st.colIdx === idx && st.dir === 'desc') ? 'asc' : 'desc';
          _sortState[id] = {colIdx: idx, dir: dir};
          _sortTable(table, ths, idx, dir);
        });
      });
    }

    function _initAll() {
      doc.querySelectorAll('.ct-tbl').forEach(_initTableSort);
    }
    _initAll();

    if (!win.__ctMutObs) {
      win.__ctMutObs = new MutationObserver(_initAll);
      win.__ctMutObs.observe(doc.body, {childList: true, subtree: true});
    }

    function _addShiftScroll(el) {
      if (el.__ctShiftScroll) return;
      el.__ctShiftScroll = true;
      el.addEventListener('wheel', function(e) {
        if (!e.shiftKey) return;
        e.preventDefault();
        el.scrollLeft += e.deltaY * 1.2;
      }, { passive: false });
    }
    function _initShiftScroll() {
      doc.querySelectorAll('.ct-outer').forEach(_addShiftScroll);
    }
    _initShiftScroll();
    if (!win.__ctShiftObs) {
      win.__ctShiftObs = new MutationObserver(_initShiftScroll);
      win.__ctShiftObs.observe(doc.body, {childList: true, subtree: true});
    }

  } catch(e) {}
})();
</script>"""


def _inject_sticky_js():
    _st_comps.html(_STICKY_JS, height=1)


# -- Cell renderer -------------------------------------------------------------

def _render_cell(key, val, row_dict=None):
    """Dispatch a (column_key, value) pair to the right HTML formatter."""
    if key == "name":
        full_name = str(val or "")
        disp_name = _SHORT_NAMES.get(full_name, full_name)
        safe_full = _html_lib.escape(full_name)
        safe_disp = _html_lib.escape(disp_name)
        ticker = row_dict.get("ticker", "") if row_dict else ""
        safe_ticker = _html_lib.escape(str(ticker))
        name_span = (
            f'<span title="{safe_full}" style="font-weight:500;color:#111827;'
            f'white-space:nowrap;display:inline-block;max-width:132px;'
            f'overflow:hidden;text-overflow:ellipsis">{safe_disp}</span>'
        )
        # M&A flag
        if ticker in _MA_NOTES:
            name_span += '<sup style="color:#F59E0B;font-size:9px;font-weight:700;margin-left:2px">*</sup>'
        # Wrap in link
        if ticker:
            name_span = (
                f'<a class="ct-name-link" href="/Company?ticker={safe_ticker}" target="_self">'
                f'{name_span}</a>'
            )
        return name_span

    if key == "ticker":
        safe = _html_lib.escape(str(val or ""))
        logo = logo_img_tag(str(val or ""), size=14)
        href = f"/Company?ticker={safe}"
        logo_span = (
            f'<span style="display:inline-flex;align-items:center;'
            f'width:14px;height:14px;margin-right:3px;flex-shrink:0;">{logo}</span>'
            if logo else ""
        )
        return (
            f'<a class="ct-tkr" href="{href}" target="_self" '
            f'style="display:inline-flex;align-items:center;text-decoration:none;">'
            f'{logo_span}<span>{safe}</span></a>'
        )

    if key == "segment":
        return _cell_segment_pill(val)

    if key in ("mkt_cap_m", "tev_m"):
        return _cell_dollar_m(val)
    if key == "price":
        return _cell_price_fmt(val)
    if key == "pct_52wk":
        return _cell_52wk_html(val)
    if key in ("ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"):
        return _cell_mult(val)
    if key in ("rev_gr", "cagr_3y"):
        return _cell_rev_growth_html(val)
    if key == "gm":
        return _cell_gm_html(val)
    if key == "ebitda_mgn":
        return _cell_ebitda_mgn_html(val)
    if key in ("chg_2w", "chg_2m"):
        return _cell_price_change_html(val)
    # Fallback
    if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
        return _cell_nd()
    return _html_lib.escape(str(val))


# -- Sort value helper ---------------------------------------------------------

def _sort_val(key, val):
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    if key in _MULT_KEYS and isinstance(val, (int, float)) and val <= 0:
        return ""
    if isinstance(val, (int, float)):
        return str(float(val))
    return _html_lib.escape(str(val).lower())


# -- HTML builders -------------------------------------------------------------

def _thead_html(cfg):
    """Two-row <thead>: group super-header row + column name row."""
    cols         = cfg["cols"]
    groups       = cfg["groups"]
    group_starts = cfg["group_starts"]

    gr_cells = []
    for gname, colspan, bg, fg, border_top in groups:
        border_css = f"border-top:3px solid {border_top};" if gname else ""
        label = _html_lib.escape(gname)
        gr_cells.append(
            f'<th class="ct-gr" colspan="{colspan}" '
            f'style="background:{bg};color:{fg};{border_css}">{label}</th>'
        )

    col_cells = []
    for i, (label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        gs_cls  = " ct-gs"  if i in group_starts else ""
        ltm_cls = " ct-ltm" if key in _LTM_KEYS else ""
        disp   = '<br>'.join(_html_lib.escape(part) for part in label.split('\n'))
        col_cells.append(
            f'<th class="ct-ch{sticky_cls}{gs_cls}{ltm_cls}" '
            f'data-col="{_html_lib.escape(key)}" '
            f'style="min-width:{width}px;{left_style}">'
            f'{disp}</th>'
        )

    return (
        '<thead class="ct-thead">'
        f"<tr>{''.join(gr_cells)}</tr>"
        f"<tr>{''.join(col_cells)}</tr>"
        "</thead>"
    )


def _data_row_html(row_dict, idx, cfg):
    """Return a <tr> for one company data row."""
    cols         = cfg["cols"]
    group_starts = cfg["group_starts"]
    alt_cls = " ct-alt" if idx % 2 == 1 else ""
    cells = []
    for i, (label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        align_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
        gs_cls    = " ct-gs"  if i in group_starts else ""
        ltm_cls   = " ct-ltm" if key in _LTM_KEYS  else ""
        val       = row_dict.get(key)
        content   = _render_cell(key, val, row_dict)
        sort_v    = _sort_val(key, val)
        cells.append(
            f'<td class="ct-td{sticky_cls}{align_cls}{gs_cls}{ltm_cls}" '
            f'data-sort="{sort_v}" style="{left_style}">'
            f'{content}</td>'
        )
    return f'<tr class="ct-tr{alt_cls}">{"".join(cells)}</tr>'


def _summary_row_html(vals_d, label, cfg, extra_cls=""):
    """Return a <tr> for a Mean or Median summary row."""
    cols         = cfg["cols"]
    group_starts = cfg["group_starts"]
    cells = []
    for i, (col_label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        align_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
        gs_cls    = " ct-gs"  if i in group_starts else ""
        ltm_cls   = " ct-ltm" if key in _LTM_KEYS  else ""

        if key == "name":
            content = (
                f'<span style="color:#94A3B8;font-style:italic;'
                f'font-size:11px;font-weight:400">{_html_lib.escape(label)}</span>'
            )
        elif key in ("ticker", "mkt_cap_m", "tev_m", "segment"):
            content = ""
        else:
            content = _render_cell(key, vals_d.get(key))

        cells.append(
            f'<td class="ct-td{sticky_cls}{align_cls}{gs_cls}{ltm_cls}" style="{left_style}">'
            f'{content}</td>'
        )
    return f'<tr class="ct-tr ct-sum{extra_cls}">{"".join(cells)}</tr>'


def _build_table_html(df, mean_d, median_d, cfg):
    """Assemble complete table HTML (no max-height scroll constraint)."""
    parts = [
        _CT_CSS,
        '<div class="ct-outer"><table class="ct-tbl">',
        _thead_html(cfg),
        "<tbody>",
        _summary_row_html(mean_d,   "Mean",   cfg, " ct-mean"),
        _summary_row_html(median_d, "Median", cfg, " ct-med"),
    ]
    for idx, row_dict in enumerate(df.to_dict("records")):
        parts.append(_data_row_html(row_dict, idx, cfg))
    parts.append("</tbody></table></div>")
    return "".join(parts)


# -- Excel export --------------------------------------------------------------

_EXCEL_COLUMNS = [
    ("Company",          "name",          28, None),
    ("Ticker",           "ticker",        10, None),
    ("Segment",          "segment",       16, None),
    ("Mkt Cap ($mm)",    "mkt_cap_m",     14, '#,##0;(#,##0)'),
    ("TEV ($mm)",        "tev_m",         14, '#,##0;(#,##0)'),
    ("% 52W Hi",         "pct_52wk",      10, '0%;(0%)'),
    ("NTM EV/Rev",       "ev_rev",        11, '0.0"x";(0.0"x")'),
    ("LTM EV/Rev",       "ltm_rev_x",     11, '0.0"x";(0.0"x")'),
    ("NTM EV/GP",        "ev_gp",         11, '0.0"x";(0.0"x")'),
    ("LTM EV/GP",        "ltm_gp_x",      11, '0.0"x";(0.0"x")'),
    ("NTM EV/EBITDA",    "ev_ebitda",     13, '0.0"x";(0.0"x")'),
    ("LTM EV/EBITDA",    "ltm_ebitda_x",  13, '0.0"x";(0.0"x")'),
    ("NTM Rev Gr%",      "rev_gr",        12, '0.0%;(0.0%)'),
    ("3Y CAGR",          "cagr_3y",       10, '0.0%;(0.0%)'),
    ("Gross Mgn",        "gm",            10, '0%;(0%)'),
    ("NTM EBITDA Mgn",   "ebitda_mgn",    14, '0%;(0%)'),
    ("1M Chg",           "chg_2w",        10, '0.0%;(0.0%)'),
    ("1Y Chg",           "chg_2m",        10, '0.0%;(0.0%)'),
]


def _build_filtered_excel(df, mean_d, median_d):
    """Build an Excel workbook from the currently filtered comps view."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    wb = Workbook()
    ws = wb.active
    ws.title = "Comps Table"

    columns = _EXCEL_COLUMNS

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sum_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    sum_font = Font(name="Calibri", bold=True, size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font_black = Font(name="Calibri", size=10, color="111827")
    data_font_blue = Font(name="Calibri", size=10, color="0000FF")
    data_align_r = Alignment(horizontal="right", vertical="center")
    data_align_l = Alignment(horizontal="left", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))
    thick_border = Border(bottom=Side(style="medium", color="334155"))

    # Title
    ws.cell(row=1, column=1, value="Healthcare Comps Table")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Data as of {date.today().strftime('%B %d, %Y')}. All financials in $mm.")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="6B7280")

    # Headers (row 4)
    hdr_row = 4
    for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thick_border
        ws.column_dimensions[cell.column_letter].width = width

    # Summary rows (Mean / Median)
    def _write_summary_row(row_num, label, summary_dict):
        for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = sum_fill
            cell.font = sum_font
            cell.border = thin_border
            cell.alignment = data_align_r
            if col_idx == 1:
                cell.value = label
                cell.alignment = data_align_l
            elif col_idx <= 3:
                continue
            else:
                val = summary_dict.get(key)
                if val is not None and not (isinstance(val, float) and (val != val)):
                    cell.value = val
                    if fmt:
                        cell.number_format = fmt

    _write_summary_row(hdr_row + 1, "Mean", mean_d)
    _write_summary_row(hdr_row + 2, "Median", median_d)

    # Data rows
    data_start = hdr_row + 3
    for i, (_, row) in enumerate(df.iterrows()):
        row_num = data_start + i
        is_alt = i % 2 == 1
        for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

            val = row.get(key)
            # Map segment key to display name for Excel
            if key == "segment" and val:
                val = SEGMENT_SHORT.get(str(val), str(val))

            if val is None or (isinstance(val, float) and val != val):
                cell.font = data_font_black
                cell.value = None
                continue

            cell.value = val
            if col_idx <= 3:
                cell.font = data_font_black
                cell.alignment = data_align_l
            else:
                cell.font = data_font_blue
                cell.alignment = data_align_r
                if fmt:
                    cell.number_format = fmt

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# -- Render --------------------------------------------------------------------

_inject_sticky_js()

# Download Excel button + subtitle
dl_col, spacer_col = st.columns([1, 5])
with dl_col:
    xlsx_data = _build_filtered_excel(df, mean_d, median_d)
    st.download_button(
        label="Download Excel",
        data=xlsx_data,
        file_name="healthcare_comps_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.markdown(
    '<p style="color:#94A3B8;font-size:11px;margin:0 0 2px 0;">'
    "Sorted by TEV descending. Click any column header to re-sort.</p>",
    unsafe_allow_html=True,
)

# Render the table
table_html = _build_table_html(df, mean_d, median_d, _CFG)
st.markdown(table_html, unsafe_allow_html=True)

# -- M&A footnote --------------------------------------------------------------
if "ticker" in df.columns:
    present = [(t, _MA_NOTES[t]) for t in df["ticker"] if t in _MA_NOTES]
    if present:
        lines = "<br>".join(
            f'<b>* {t}</b> &mdash; {note}' for t, note in present
        )
        st.markdown(
            f'<p style="color:#6B7280;font-size:10.5px;margin-top:2px;'
            f'margin-bottom:12px;line-height:1.6">{lines}</p>',
            unsafe_allow_html=True,
        )

# -- Data source attribution ---------------------------------------------------

try:
    last_fetch = db.get_last_fetch_time()
    if last_fetch:
        dt = datetime.strptime(last_fetch[:10], "%Y-%m-%d")
        date_str = dt.strftime("%B %d, %Y")
    else:
        date_str = "N/A"
except Exception:
    date_str = "N/A"

st.markdown(
    f'<p style="color:#94A3B8;font-size:11px;margin-top:6px;margin-bottom:0;">'
    f'Data: FactSet (fundamentals & estimates) &middot; As of {date_str}</p>',
    unsafe_allow_html=True,
)
