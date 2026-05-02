"""
Comp table renderer — custom HTML table, Meritech-style layout.
Sticky Company + Ticker columns. Mean / Median summary rows pinned at top.
Light theme.
"""

import html as _html_lib
import io
import json
import streamlit as st
import streamlit.components.v1 as _st_comps
import numpy as np
from pathlib import Path

from components.formatters import (
    COMPS_NUMERIC_COLS,
    build_comps_df,
    compute_comps_summary,
    _cell_dollar_m,
    _cell_price_fmt,
    _cell_mult,
    _cell_ga_mult,
    _cell_52wk_html,
    _cell_rev_growth_html,
    _cell_gm_html,
    _cell_ebitda_mgn_html,
    _cell_price_change_html,
    _cell_nd,
)
from config.settings import SUB_SEGMENT_DISPLAY, EXCEL_OVERRIDE_PATH
from fetcher.excel_override import load_overrides, apply_overrides
from components.logos import logo_img_tag

# ── Column definitions ────────────────────────────────────────────────────────
# (header_label, df_key, min_width_px, is_sticky, text_align)

# Revenue view — NTM + LTM valuation multiples (default, 24 columns)
# Two-line labels: "\n" is rendered as <br> in _thead_html().
# All non-sticky data columns use a uniform 76px min-width (Meritech-style).
_COLS_REVENUE = [
    # Info group (sticky)
    ("Company",           "name",          160, True,  "left"),
    ("Ticker",            "ticker",         70, True,  "left"),
    # Market Data
    ("Mkt\nCap",          "mkt_cap_m",      76, False, "right"),
    ("TEV",               "tev_m",          76, False, "right"),
    ("%\n52W Hi",         "pct_52wk",       76, False, "right"),
    # NTM Multiples (Rev → GP → EBITDA)
    ("NTM\nEV/Rev",       "ev_rev",         76, False, "right"),
    ("NTM\nEV/GP",        "ev_gp",          76, False, "right"),
    ("NTM\nEV/EBITDA",    "ev_ebitda",      76, False, "right"),
    # LTM Multiples (Rev → GP → EBITDA)
    ("LTM\nEV/Rev",       "ltm_rev_x",      76, False, "right"),
    ("LTM\nEV/GP",        "ltm_gp_x",       76, False, "right"),
    ("LTM\nEV/EBITDA",    "ltm_ebitda_x",   76, False, "right"),
    # Growth & Margins
    ("NTM Rev\nGr%",      "rev_gr",          76, False, "right"),
    ("3Y\nCAGR",          "cagr_3y",         76, False, "right"),
    ("Gross\nMgn",        "gm",              76, False, "right"),
    ("EBITDA\nMgn",       "ebitda_mgn",      76, False, "right"),
    # Price Performance
    ("2W\nChg",           "chg_2w",          76, False, "right"),
    ("2M\nChg",           "chg_2m",          76, False, "right"),
]

_GROUPS_REVENUE = [
    ("",                   2, "#1E293B",     "transparent", "transparent"),
    ("Market Data",        3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("NTM Multiples",      3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("LTM Multiples",      3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Growth & Margins",   4, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Price Performance",  2, "#1E293B",     "#FFFFFF",     "transparent"),
]

# Col indices where a group-divider left-border is applied
# 0:Company 1:Ticker | 2:MktCap 3:TEV 4:52WH | 5:NTM EV/Rev 6:EV/GP 7:EV/EBITDA
# 8:LTM EV/Rev 9:EV/GP 10:EV/EBITDA | 11:NTM Rev Gr% 12:3Y CAGR 13:GM 14:EBITDA Mgn | 15:2W Chg 16:2M Chg
_GROUP_STARTS_REVENUE = {2, 5, 8, 11, 15}

# EBITDA view — NTM + LTM EV/EBITDA only (20 columns)
# Same two-line label convention; uniform 76px non-sticky columns.
_COLS_EBITDA = [
    ("Company",           "name",          160, True,  "left"),
    ("Ticker",            "ticker",         70, True,  "left"),
    ("Mkt\nCap",          "mkt_cap_m",      76, False, "right"),
    ("TEV",               "tev_m",          76, False, "right"),
    ("%\n52W Hi",         "pct_52wk",       76, False, "right"),
    # EBITDA multiples only
    ("NTM\nEV/EBITDA",    "ev_ebitda",      76, False, "right"),
    ("LTM\nEV/EBITDA",    "ltm_ebitda_x",   76, False, "right"),
    # Growth & Margins
    ("NTM Rev\nGr%",      "rev_gr",          76, False, "right"),
    ("3Y\nCAGR",          "cagr_3y",         76, False, "right"),
    ("Gross\nMgn",        "gm",              76, False, "right"),
    ("EBITDA\nMgn",       "ebitda_mgn",      76, False, "right"),
    # Price Performance
    ("2W\nChg",           "chg_2w",          76, False, "right"),
    ("2M\nChg",           "chg_2m",          76, False, "right"),
]

_GROUPS_EBITDA = [
    ("",                   2, "#1E293B",     "transparent", "transparent"),
    ("Market Data",        3, "#1E293B",     "#FFFFFF",     "transparent"),
    ("NTM Multiples",      1, "#1E293B",     "#FFFFFF",     "transparent"),
    ("LTM Multiples",      1, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Growth & Margins",   4, "#1E293B",     "#FFFFFF",     "transparent"),
    ("Price Performance",  2, "#1E293B",     "#FFFFFF",     "transparent"),
]

# 0:Company 1:Ticker | 2:MktCap 3:TEV 4:52WH | 5:NTM EV/EBITDA | 6:LTM EV/EBITDA
# 7:NTM Rev Gr% 8:3Y CAGR 9:GM 10:EBITDA Mgn | 11:2W Chg 12:2M Chg
_GROUP_STARTS_EBITDA = {2, 5, 6, 7, 11}

# ── View config dicts ─────────────────────────────────────────────────────────
_CFG_REVENUE = {
    "cols":         _COLS_REVENUE,
    "groups":       _GROUPS_REVENUE,
    "group_starts": _GROUP_STARTS_REVENUE,
    "tev_idx":      3,
}
_CFG_EBITDA = {
    "cols":         _COLS_EBITDA,
    "groups":       _GROUPS_EBITDA,
    "group_starts": _GROUP_STARTS_EBITDA,
    "tev_idx":      3,
}

# Backwards-compat alias (not used by HTML builders)
_COLS = _COLS_REVENUE

# ── LTM column keys — receive lighter text styling via ct-ltm CSS class ──────
_LTM_KEYS = {"ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}

# ── Multiple column keys — non-positive values sort last (N/M treatment) ──────
_MULT_KEYS = {"ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"}

# ── Short display names (full DB name → condensed label, full name stays as tooltip) ──
_SHORT_NAMES = {
    "Zoom Video Communications, Inc.": "Zoom",
    "Zoom Video Communications":       "Zoom",
    "Salesforce, Inc.":                "Salesforce",
    "ServiceNow, Inc.":                "ServiceNow",
    "Microsoft Corporation":           "Microsoft",
    "Alphabet Inc.":                   "Alphabet",
    "Workday, Inc.":                   "Workday",
    "HubSpot, Inc.":                   "HubSpot",
    "Datadog, Inc.":                   "Datadog",
    "Snowflake Inc.":                  "Snowflake",
    "MongoDB, Inc.":                   "MongoDB",
    "Palantir Technologies Inc.":      "Palantir",
    "UiPath Inc.":                     "UiPath",
    "AppFolio, Inc.":                  "AppFolio",
    "Veeva Systems Inc.":              "Veeva",
    "Procore Technologies, Inc.":      "Procore",
    "Twilio Inc.":                     "Twilio",
    "Okta, Inc.":                      "Okta",
    "CrowdStrike Holdings, Inc.":      "CrowdStrike",
    "Palo Alto Networks, Inc.":        "Palo Alto Networks",
    "SentinelOne, Inc.":               "SentinelOne",
    "Zscaler, Inc.":                   "Zscaler",
    "Elastic N.V.":                    "Elastic",
    "Confluent, Inc.":                 "Confluent",
    "Amplitude, Inc.":                 "Amplitude",
    "Braze, Inc.":                     "Braze",
    "HashiCorp, Inc.":                 "HashiCorp",
    "Sprinklr, Inc.":                  "Sprinklr",
    "Yext, Inc.":                      "Yext",
    "RingCentral, Inc.":               "RingCentral",
    "Five9, Inc.":                     "Five9",
    "8x8, Inc.":                       "8x8",
    "ZoomInfo Technologies Inc.":      "ZoomInfo",
    "Clearbit, Inc.":                  "Clearbit",
    "Coupa Software Incorporated":     "Coupa",
    "Asana, Inc.":                     "Asana",
    "Notion Labs, Inc.":               "Notion",
    "Carta, Inc.":                     "Carta",
    "Rippling Technologies Inc.":      "Rippling",
    "Airtable, Inc.":                  "Airtable",
    "DigitalBridge Group, Inc.":       "DigitalBridge",
    "monday.com Ltd.":                 "Monday.com",
    "monday.com":                      "Monday.com",
    "Shopify Inc.":                    "Shopify",
    "Cloudflare, Inc.":                "Cloudflare",
    "Fastly, Inc.":                    "Fastly",
    "PagerDuty, Inc.":                 "PagerDuty",
    "Dynatrace, Inc.":                 "Dynatrace",
    "New Relic, Inc.":                 "New Relic",
    "Sumo Logic, Inc.":                "Sumo Logic",
    "Splunk Inc.":                     "Splunk",
    "Informatica Inc.":                "Informatica",
    "Alteryx, Inc.":                   "Alteryx",
    "MicroStrategy Incorporated":      "MicroStrategy",
    "Verint Systems Inc.":             "Verint",
    "Sprout Social, Inc.":             "Sprout Social",
    "Zendesk, Inc.":                   "Zendesk",
    "Freshworks Inc.":                 "Freshworks",
    "Zenvia Inc.":                     "Zenvia",
    "Bandwidth Inc.":                  "Bandwidth",
    "LivePerson, Inc.":                "LivePerson",
    "Zuora, Inc.":                     "Zuora",
    "Recurly, Inc.":                   "Recurly",
    "Chargebee Technologies Inc.":     "Chargebee",
    "Paddle.com, Inc.":                "Paddle",
    "Maxio Inc.":                      "Maxio",
    "ClickUp Technologies, Inc.":      "ClickUp",
    "Wrike, Inc.":                     "Wrike",
    "Smartsheet Inc.":                 "Smartsheet",
    "Agiloft, Inc.":                   "Agiloft",
    "Ironclad, Inc.":                  "Ironclad",
    "ContractPodAi Ltd.":              "ContractPodAi",
    "DocuSign, Inc.":                  "DocuSign",
    "Adobe Inc.":                      "Adobe",
    "Autodesk, Inc.":                  "Autodesk",
    "Bentley Systems, Incorporated":   "Bentley Systems",
    "Trimble Inc.":                    "Trimble",
    "Instructure Holdings, Inc.":      "Instructure",
    "Blackbaud, Inc.":                 "Blackbaud",
    "Toast, Inc.":                     "Toast",
    "Lightspeed Commerce Inc.":        "Lightspeed",
    "PAX Global Technology Limited":   "PAX Global",
    "NCR Voyix Corporation":           "NCR Voyix",
    "NCR Corporation":                 "NCR",
    "Roper Technologies, Inc.":        "Roper Tech",
    "Tyler Technologies, Inc.":        "Tyler Tech",
    "Veritiv Corporation":             "Veritiv",
    "Rapid7, Inc.":                    "Rapid7",
    "Tenable Holdings, Inc.":          "Tenable",
    "Qualys, Inc.":                    "Qualys",
    "Varonis Systems, Inc.":           "Varonis",
    "Darktrace plc":                   "Darktrace",
    "Vectra AI, Inc.":                 "Vectra AI",
    "Exabeam, Inc.":                   "Exabeam",
    "Abnormal Security Corporation":   "Abnormal Security",
    "Proofpoint, Inc.":                "Proofpoint",
    "Mimecast Limited":                "Mimecast",
    "KnowBe4, Inc.":                   "KnowBe4",
    "OneTrust, LLC":                   "OneTrust",
    "BigBear.ai Holdings, Inc.":       "BigBear.ai",
    "C3.ai, Inc.":                     "C3.ai",
    "Samsara Inc.":                    "Samsara",
    "Domo, Inc.":                      "Domo",
    "Qlik Technologies Inc.":          "Qlik",
    "ThoughtSpot Inc.":                "ThoughtSpot",
    "Sisense Ltd.":                    "Sisense",
    "Sigma Computing, Inc.":           "Sigma",
    "OutSystems, S.A.":                "OutSystems",
    "Appian Corporation":              "Appian",
    "Mendix Technology B.V.":          "Mendix",
    "Creatio":                         "Creatio",
    "Salesforce.com, Inc.":            "Salesforce",
    "Oracle Corporation":              "Oracle",
    "SAP SE":                          "SAP",
    "Workiva Inc.":                    "Workiva",
    "Netsol Technologies, Inc.":       "Netsol",
    "nCino, Inc.":                     "nCino",
    "Temenos AG":                      "Temenos",
    "Finastra Group Holdings Limited": "Finastra",
    "Q2 Holdings, Inc.":               "Q2",
    "Blend Labs, Inc.":                "Blend Labs",
    "Meridian Link Holding Corp.":     "MeridianLink",
    "OS Inc.":                         "OpenText",
    "Open Text Corporation":           "OpenText",
    "Atlassian Corporation":           "Atlassian",
    "GitLab Inc.":                     "GitLab",
    "GitHub, Inc.":                    "GitHub",
    "JFrog Ltd.":                      "JFrog",
    "Gradle Enterprise, LLC":          "Gradle",
    "CircleCI, Inc.":                  "CircleCI",
    "LaunchDarkly":                    "LaunchDarkly",
    "Split Software, Inc.":            "Split",
    "Harness Inc.":                    "Harness",
    "Honeycomb.io, Inc.":              "Honeycomb",
    "Grafana Labs":                    "Grafana",
}

# ── Pending / rumored M&A annotations (ticker → footnote text) ───────────────
_MA_NOTES = {
    "OS":   "Pending take-private — Hg Capital ($6.4B all-cash; expected H1 2026)",
}

_COL1_W = 160   # Company column pixel width (left offset for Ticker sticky)

# ── Background colors (opaque — sticky cells cannot use rgba) ─────────────────
_BG_BASE  = "#FFFFFF"
_BG_ALT   = "#FAFBFD"   # barely-there tint — 1% gray
_BG_HOVER = "#F0F4FF"   # soft blue on hover
_BG_SUM   = "#E2E8F0"   # light slate for mean/median summary rows

# ── CSS ───────────────────────────────────────────────────────────────────────
_CT_CSS = """<style>
.ct-outer {
    overflow: auto;
    max-height: calc(100vh - 130px);
    -webkit-overflow-scrolling: touch;
    border: 1px solid #D1D5DB;
    border-radius: 4px;
    margin-bottom: 4px;
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, 'Inter', 'Segoe UI', sans-serif;
    background: %(base)s;
}
.ct-tbl {
    border-collapse: separate;
    border-spacing: 0;
    min-width: 1610px;
    width: 100%%;
    font-size: 12px;
    font-variant-numeric: tabular-nums;
    background: %(base)s;
}
/* ── Sticky thead — CSS position:sticky works because .ct-outer scrolls
       both axes (overflow:auto), so the scroll parent IS the container ── */
.ct-thead { position: sticky; top: 0; z-index: 22; background: #1E293B; }
/* ── Group header row — colored band above each column group ── */
.ct-gr {
    padding: 6px 10px;
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.07em;
    text-transform: uppercase;
    text-align: center;
    white-space: nowrap;
}
/* ── Column name row — two-line headers, vertically bottom-aligned ── */
.ct-ch {
    padding: 10px 6px 8px;
    font-size: 10.5px;
    font-weight: 700;
    color: #FFFFFF;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    background: #1E293B;
    text-align: center;
    border-bottom: none;
    white-space: normal;
    line-height: 1.4;
    vertical-align: bottom;
}
.ct-ch.lft { text-align: left; }
.ct-ch.ctr { text-align: center; }
.ct-ch.ct-s0, .ct-ch.ct-s1 { z-index: 26; background: #1E293B; }
/* ── Group divider (left border on first col of each group) ── */
.ct-gs { border-left: 1px solid #CBD5E1; }
/* ── In the header rows, use inset box-shadow instead of border (no gaps) ── */
.ct-thead .ct-gs { border-left: none; box-shadow: inset 1px 0 0 #334155; }
/* ── Ensure ALL header cells are fully opaque (no bleed-through on scroll) ── */
.ct-gr, .ct-ch { background: #1E293B !important; }
.ct-thead tr { background: #1E293B !important; }
.ct-thead th { background: #1E293B !important; }
/* ── Data cells — tighter rows, softer gray ── */
.ct-td {
    padding: 2px 8px;
    text-align: right;
    border-bottom: 1px solid #F3F4F6;
    background: %(base)s;
    vertical-align: middle;
    white-space: nowrap;
    color: #111827;
    line-height: 1.2;
}
.ct-td.lft { text-align: left; }
.ct-td.ctr { text-align: center; }
.ct-tr:hover .ct-td      { background-color: %(hover)s !important; }
.ct-alt .ct-td            { background: %(alt)s; }
.ct-alt:hover .ct-td      { background-color: %(hover)s !important; }
/* ── Summary rows — subtle tint + borders provide visual distinction ── */
.ct-sum .ct-td {
    background: %(sum)s !important;
    font-weight: 600;
    font-size: 12px;
    padding: 2px 8px;
    border-bottom: 1px solid #E2E8F0;
    color: #111827;
}
.ct-sum .ct-td span { color: #111827 !important; }
.ct-mean .ct-td { border-top: 1px solid #94A3B8 !important; }
.ct-med .ct-td  { border-bottom: 2px solid #94A3B8 !important; }
/* ── LTM column cells — muted slate tone to distinguish from NTM ── */
.ct-ltm.ct-td { color: #111827; }
.ct-ltm.ct-ch { color: #FFFFFF; }
/* ── Sticky columns ── */
.ct-s0 { position: sticky; left: 0;        z-index: 6; }
.ct-s1 { position: sticky; left: 160px;    z-index: 6; border-right: 1px solid #E5E7EB; }
/* Sticky cells match their row's background */
.ct-alt .ct-td.ct-s0,
.ct-alt .ct-td.ct-s1       { background: %(alt)s; }
.ct-sum .ct-td.ct-s0,
.ct-sum .ct-td.ct-s1       { background: %(sum)s !important; }
.ct-tr:hover .ct-td.ct-s0,
.ct-tr:hover .ct-td.ct-s1,
.ct-alt:hover .ct-td.ct-s0,
.ct-alt:hover .ct-td.ct-s1 { background-color: %(hover)s !important; }
/* ── Ticker link — blue accent, medium weight ── */
.ct-tkr { color:#111827; font-weight:600; font-size:11px; text-decoration:none; }
.ct-tkr:hover { text-decoration:underline; }
/* ── Sortable column headers ── */
.ct-ch[data-col] { cursor: pointer; user-select: none; }
.ct-ch[data-col]:hover { color: #FFFFFF; }
.ct-ch[data-col]::after {
    content: "↕";
    color: #D1D5DB;
    display: block;
    font-size: 8px;
    line-height: 1;
    margin-top: 2px;
    text-align: center;
}
.ct-ch.ct-sort-asc  { color: #3B82F6 !important; }
.ct-ch.ct-sort-desc { color: #3B82F6 !important; }
.ct-ch.ct-sort-asc::after  { content: "▲"; color: #3B82F6; display: block; font-size: 8px; line-height: 1; margin-top: 2px; text-align: center; }
.ct-ch.ct-sort-desc::after { content: "▼"; color: #3B82F6; display: block; font-size: 8px; line-height: 1; margin-top: 2px; text-align: center; }
</style>""" % {"base": _BG_BASE, "alt": _BG_ALT, "hover": _BG_HOVER, "sum": _BG_SUM}


# ── Sortable-column + shift-scroll JS (injected once per page) ───────────────
# Sticky header is handled purely by CSS (position:sticky on .ct-thead) since
# .ct-outer now has overflow:auto, making it the scroll parent for both axes.
_STICKY_JS = """<script>
(function(){
  try {
    var win = window.parent;
    var doc = win.document;

    /* ── Sortable columns ──────────────────────────────────────────────── */
    if (!win.__ctSortState) win.__ctSortState = {};
    var _sortState = win.__ctSortState;

    function _cmp(aSort, bSort, dir) {
      if (aSort === '' && bSort === '') return 0;
      if (aSort === '') return 1;   /* N/A always last */
      if (bSort === '') return -1;
      var an = parseFloat(aSort), bn = parseFloat(bSort);
      var v  = (isNaN(an) || isNaN(bn)) ? aSort.localeCompare(bSort) : an - bn;
      return dir === 'asc' ? v : -v;
    }

    function _sortTable(table, ths, colIdx, dir) {
      var tbody = table.querySelector('tbody');
      if (!tbody) return;
      var all  = Array.from(tbody.querySelectorAll('tr'));
      var summ = all.filter(function(r){ return r.classList.contains('ct-sum'); });
      var data = all.filter(function(r){ return !r.classList.contains('ct-sum'); });
      data.sort(function(a, b) {
        var tds = function(r){ return r.querySelectorAll('td'); };
        var ac  = tds(a)[colIdx], bc = tds(b)[colIdx];
        return _cmp(
          ac ? (ac.dataset.sort || '') : '',
          bc ? (bc.dataset.sort || '') : '',
          dir
        );
      });
      /* Re-append: summary rows pinned first, then sorted data */
      summ.concat(data).forEach(function(r){ tbody.appendChild(r); });
      /* Re-stripe alternating rows */
      data.forEach(function(r, i){
        r.classList.toggle('ct-alt', i % 2 === 1);
      });
      /* Update sort icons */
      ths.forEach(function(th, i){
        th.classList.toggle('ct-sort-asc',  i === colIdx && dir === 'asc');
        th.classList.toggle('ct-sort-desc', i === colIdx && dir === 'desc');
      });
    }

    function _initTableSort(table) {
      if (table.__ctSortBound) return;
      table.__ctSortBound = true;
      var id = 'ct_' + Math.random().toString(36).substr(2,9);
      table.__ctId = id;
      _sortState[id] = {colIdx: -1, dir: 'desc'};
      var thead = table.querySelector('thead.ct-thead');
      if (!thead) return;
      var hrows = thead.querySelectorAll('tr');
      if (hrows.length < 2) return;
      var ths = Array.from(hrows[1].querySelectorAll('th[data-col]'));
      ths.forEach(function(th, idx){
        th.addEventListener('click', function(){
          var st  = _sortState[id];
          var dir = (st.colIdx === idx && st.dir === 'desc') ? 'asc' : 'desc';
          _sortState[id] = {colIdx: idx, dir: dir};
          _sortTable(table, ths, idx, dir);
        });
      });
    }

    function _initAll() {
      doc.querySelectorAll('.ct-tbl').forEach(_initTableSort);
    }
    _initAll();

    /* Watch for Streamlit re-renders adding new tables */
    if (!win.__ctMutObs) {
      win.__ctMutObs = new MutationObserver(_initAll);
      win.__ctMutObs.observe(doc.body, {childList: true, subtree: true});
    }

    /* ── Shift+scroll → horizontal pan ─────────────────────────────── */
    function _addShiftScroll(el) {
      if (el.__ctShiftScroll) return;
      el.__ctShiftScroll = true;
      el.addEventListener('wheel', function(e) {
        if (!e.shiftKey) return;
        e.preventDefault();
        el.scrollLeft += e.deltaY * 1.2;
      }, { passive: false });
    }
    function _initShiftScroll() {
      doc.querySelectorAll('.ct-outer').forEach(_addShiftScroll);
    }
    _initShiftScroll();
    if (!win.__ctShiftObs) {
      win.__ctShiftObs = new MutationObserver(_initShiftScroll);
      win.__ctShiftObs.observe(doc.body, {childList: true, subtree: true});
    }

  } catch(e) {}
})();
</script>"""


def _inject_sticky_js():
    """Render a 1px-height component that executes the sticky-header rAF loop.
    height=1 (not 0) is required — height=0 can suppress iframe JS execution."""
    _st_comps.html(_STICKY_JS, height=1)


# ── NRR helpers ───────────────────────────────────────────────────────────────

_NRR_JSON_PATH = Path(__file__).parent.parent / "data" / "nrr_data.json"


@st.cache_data(ttl=3600)
def _load_nrr_data():
    """Stubbed: NRR/GRR data is not used in healthcare comps.
    Returns empty maps so existing call sites continue to work."""
    return {}, {}, {}


def _inject_nrr(df, nrr_map, grr_map=None):
    """No-op shim — preserved so call sites need no changes."""
    return df


def _render_nrr_footnote(meta):
    """Render NRR data-source footnote below the comp table."""
    last_updated = (meta or {}).get("last_updated", "")
    suffix = f" Updated {last_updated}." if last_updated else ""
    st.markdown(
        f'<p style="color:#9CA3AF;font-size:10px;margin-top:2px;margin-bottom:10px;">'
        f'NRR — Net Revenue Retention from most recently reported earnings. '
        f'Source: Clouded Judgement / Meritech Analytics.{suffix}</p>',
        unsafe_allow_html=True,
    )


# ── Cell renderer ──────────────────────────────────────────────────────────────

def _render_cell(key, val):
    """Dispatch a (column_key, value) pair to the right HTML formatter."""
    if key == "name":
        full_name = str(val or "")
        disp_name = _SHORT_NAMES.get(full_name, full_name)
        safe_full = _html_lib.escape(full_name)
        safe_disp = _html_lib.escape(disp_name)
        return (
            f'<span title="{safe_full}" style="font-weight:500;color:#111827;'
            f'white-space:nowrap;display:inline-block;max-width:152px;'
            f'overflow:hidden;text-overflow:ellipsis">{safe_disp}</span>'
        )
    if key == "ticker":
        safe = _html_lib.escape(str(val or ""))
        logo = logo_img_tag(str(val or ""), size=14)
        href = f"/Company?ticker={safe}"
        # Logo + ticker text in a single inline-flex link so they stay together,
        # and clicking either piece routes to the in-app Company Profile page.
        logo_span = (
            f'<span style="display:inline-flex;align-items:center;'
            f'width:14px;height:14px;margin-right:4px;flex-shrink:0;">{logo}</span>'
            if logo else ""
        )
        return (
            f'<a class="ct-tkr" href="{href}" target="_self" '
            f'style="display:inline-flex;align-items:center;text-decoration:none;">'
            f'{logo_span}<span>{safe}</span></a>'
        )
    if key in ("mkt_cap_m", "tev_m"):
        return _cell_dollar_m(val)
    if key == "price":
        return _cell_price_fmt(val)
    if key == "pct_52wk":
        return _cell_52wk_html(val)
    if key in ("ev_rev", "ev_ebitda", "ev_gp", "ltm_rev_x", "ltm_ebitda_x", "ltm_gp_x"):
        return _cell_mult(val)
    if key in ("ga_rev", "ga_gp"):
        return _cell_ga_mult(val)
    if key in ("rev_gr", "cagr_3y"):
        return _cell_rev_growth_html(val)
    if key == "gm":
        return _cell_gm_html(val)
    if key == "ebitda_mgn":
        return _cell_ebitda_mgn_html(val)
    if key in ("chg_2w", "chg_2m"):
        return _cell_price_change_html(val)
    # Fallback
    if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
        return _cell_nd()
    return _html_lib.escape(str(val))


# ── HTML builders ──────────────────────────────────────────────────────────────

def _thead_html(cfg):
    """Two-row <thead>: group super-header row + column name row."""
    cols         = cfg["cols"]
    groups       = cfg["groups"]
    group_starts = cfg["group_starts"]
    tev_idx      = cfg.get("tev_idx", 3)

    # Row 1 — group spans
    gr_cells = []
    for gname, colspan, bg, fg, border_top in groups:
        border_css = f"border-top:3px solid {border_top};" if gname else ""
        label = _html_lib.escape(gname)
        gr_cells.append(
            f'<th class="ct-gr" colspan="{colspan}" '
            f'style="background:{bg};color:{fg};{border_css}">{label}</th>'
        )

    # Row 2 — individual column headers (sticky top:0)
    col_cells = []
    for i, (label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        align_cls  = ""  # All column headers center-aligned via .ct-ch default
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        gs_cls  = " ct-gs"  if i in group_starts else ""
        ltm_cls = " ct-ltm" if key in _LTM_KEYS else ""
        # Split on \n to produce two-line headers (e.g. "NTM\nEV/Rev" → "NTM<br>EV/Rev")
        disp   = '<br>'.join(_html_lib.escape(part) for part in label.split('\n'))
        col_cells.append(
            f'<th class="ct-ch{sticky_cls}{align_cls}{gs_cls}{ltm_cls}" '
            f'data-col="{_html_lib.escape(key)}" '
            f'style="min-width:{width}px;{left_style}">'
            f'{disp}</th>'
        )

    return (
        '<thead class="ct-thead">'
        f"<tr>{''.join(gr_cells)}</tr>"
        f"<tr>{''.join(col_cells)}</tr>"
        "</thead>"
    )


def _sort_val(key, val):
    """Return a string safe for a data-sort attribute (numeric or text).
    Non-positive multiples return '' so they sort last (same as N/A)."""
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    # N/M multiples (≤0) sort last — consistent with display treatment
    if key in _MULT_KEYS and isinstance(val, (int, float)) and val <= 0:
        return ""
    if isinstance(val, (int, float)):
        return str(float(val))
    return _html_lib.escape(str(val).lower())


def _data_row_html(row_dict, idx, cfg):
    """Return a <tr> for one company data row."""
    cols         = cfg["cols"]
    group_starts = cfg["group_starts"]
    alt_cls = " ct-alt" if idx % 2 == 1 else ""
    cells = []
    for i, (label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        align_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
        gs_cls    = " ct-gs"  if i in group_starts else ""
        ltm_cls   = " ct-ltm" if key in _LTM_KEYS  else ""
        val       = row_dict.get(key)
        content   = _render_cell(key, val)
        sort_v    = _sort_val(key, val)
        # Append M&A flag star to company name cell
        if key == "name" and row_dict.get("ticker") in _MA_NOTES:
            content += '<sup style="color:#F59E0B;font-size:9px;font-weight:700;margin-left:2px">*</sup>'
        # Wrap company name in link to the in-app Company Profile page
        if key == "name" and row_dict.get("ticker"):
            t = _html_lib.escape(str(row_dict["ticker"]))
            content = (
                f'<a href="/Company?ticker={t}" target="_self" '
                f'style="text-decoration:none;color:inherit;">{content}</a>'
            )
        cells.append(
            f'<td class="ct-td{sticky_cls}{align_cls}{gs_cls}{ltm_cls}" '
            f'data-sort="{sort_v}" style="{left_style}">'
            f'{content}</td>'
        )
    return f'<tr class="ct-tr{alt_cls}">{"".join(cells)}</tr>'


def _summary_row_html(vals_d, label, cfg, extra_cls=""):
    """Return a <tr> for a Mean or Median summary row."""
    cols         = cfg["cols"]
    group_starts = cfg["group_starts"]
    cells = []
    for i, (col_label, key, width, sticky, align) in enumerate(cols):
        sticky_cls = (" ct-s0" if sticky and i == 0
                      else " ct-s1" if sticky and i == 1
                      else "")
        left_style = ("left:0px;"              if i == 0 and sticky
                      else f"left:{_COL1_W}px;" if i == 1 and sticky
                      else "")
        align_cls = " lft" if align == "left" else " ctr" if align == "center" else ""
        gs_cls    = " ct-gs"  if i in group_starts else ""
        ltm_cls   = " ct-ltm" if key in _LTM_KEYS  else ""

        if key == "name":
            content = (
                f'<span style="color:#94A3B8;font-style:italic;'
                f'font-size:11px;font-weight:400">{_html_lib.escape(label)}</span>'
            )
        elif key in ("ticker", "mkt_cap_m", "tev_m"):
            content = ""
        else:
            content = _render_cell(key, vals_d.get(key))

        cells.append(
            f'<td class="ct-td{sticky_cls}{align_cls}{gs_cls}{ltm_cls}" style="{left_style}">'
            f'{content}</td>'
        )
    return f'<tr class="ct-tr ct-sum{extra_cls}">{"".join(cells)}</tr>'


def _build_table_html(df, mean_d, median_d, cfg):
    """Assemble complete table HTML (CSS + wrapper + thead + tbody)."""
    parts = [
        _CT_CSS,
        '<div class="ct-outer"><table class="ct-tbl">',
        _thead_html(cfg),
        "<tbody>",
        _summary_row_html(mean_d,   "Mean",   cfg, " ct-mean"),
        _summary_row_html(median_d, "Median", cfg, " ct-med"),
    ]
    for idx, row_dict in enumerate(df.to_dict("records")):
        parts.append(_data_row_html(row_dict, idx, cfg))
    parts.append("</tbody></table></div>")
    return "".join(parts)


# ── M&A footnote renderer ─────────────────────────────────────────────────────

def _render_ma_footnote(df):
    """If any M&A-flagged tickers appear in df, render a * footnote below the table."""
    if "ticker" not in df.columns:
        return
    present = [(t, _MA_NOTES[t]) for t in df["ticker"] if t in _MA_NOTES]
    if not present:
        return
    lines = "<br>".join(
        f'<b>* {t}</b> &mdash; {note}' for t, note in present
    )
    st.markdown(
        f'<p style="color:#6B7280;font-size:10.5px;margin-top:2px;'
        f'margin-bottom:12px;line-height:1.6">{lines}</p>',
        unsafe_allow_html=True,
    )


# ── Excel override cache ──────────────────────────────────────────────────────

@st.cache_data(ttl=60)
def _cached_load_overrides():
    return load_overrides(EXCEL_OVERRIDE_PATH)


# ── Excel export ─────────────────────────────────────────────────────────────

_EXCEL_COLUMNS = [
    ("Company",          "name",          28, None),
    ("Ticker",           "ticker",        10, None),
    ("Mkt Cap ($mm)",    "mkt_cap_m",     14, '#,##0;(#,##0)'),
    ("TEV ($mm)",        "tev_m",         14, '#,##0;(#,##0)'),
    ("% 52W Hi",         "pct_52wk",      10, '0%;(0%)'),
    ("NTM EV/Rev",       "ev_rev",        11, '0.0"x";(0.0"x")'),
    ("NTM EV/GP",        "ev_gp",         11, '0.0"x";(0.0"x")'),
    ("NTM EV/EBITDA",    "ev_ebitda",     13, '0.0"x";(0.0"x")'),
    ("LTM EV/Rev",       "ltm_rev_x",     11, '0.0"x";(0.0"x")'),
    ("LTM EV/GP",        "ltm_gp_x",      11, '0.0"x";(0.0"x")'),
    ("LTM EV/EBITDA",    "ltm_ebitda_x",  13, '0.0"x";(0.0"x")'),
    ("NTM Rev Gr%",      "rev_gr",        12, '0.0%;(0.0%)'),
    ("3Y CAGR",          "cagr_3y",       10, '0.0%;(0.0%)'),
    ("Gross Mgn",        "gm",            10, '0%;(0%)'),
    ("EBITDA Mgn",       "ebitda_mgn",    11, '0%;(0%)'),
    ("2W Chg",           "chg_2w",        10, '0.0%;(0.0%)'),
    ("2M Chg",           "chg_2m",        10, '0.0%;(0.0%)'),
]


def _write_excel_sheet(ws, df, mean_d, median_d, segment_name):
    """Write one formatted comp sheet into an existing worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    columns = _EXCEL_COLUMNS

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sum_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    sum_font = Font(name="Calibri", bold=True, size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font_black = Font(name="Calibri", size=10, color="111827")
    data_font_blue = Font(name="Calibri", size=10, color="0000FF")
    data_align_r = Alignment(horizontal="right", vertical="center")
    data_align_l = Alignment(horizontal="left", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))
    thick_border = Border(bottom=Side(style="medium", color="334155"))

    # Title
    ws.cell(row=1, column=1, value=f"{segment_name} — Comparable Companies")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Data as of {date.today().strftime('%B %d, %Y')}. All financials in $mm.")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="6B7280")

    # Headers (row 4)
    hdr_row = 4
    for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thick_border
        ws.column_dimensions[cell.column_letter].width = width

    # Summary rows (Mean / Median)
    def _write_summary_row(row_num, label, summary_dict):
        for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = sum_fill
            cell.font = sum_font
            cell.border = thin_border
            cell.alignment = data_align_r
            if col_idx == 1:
                cell.value = label
                cell.alignment = data_align_l
            elif col_idx == 2:
                continue
            else:
                val = summary_dict.get(key)
                if val is not None and not (isinstance(val, float) and (val != val)):
                    cell.value = val
                    if fmt:
                        cell.number_format = fmt

    _write_summary_row(hdr_row + 1, "Mean", mean_d)
    _write_summary_row(hdr_row + 2, "Median", median_d)

    # Data rows
    data_start = hdr_row + 3
    for i, (_, row) in enumerate(df.iterrows()):
        row_num = data_start + i
        is_alt = i % 2 == 1
        for col_idx, (header, key, width, fmt) in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

            val = row.get(key)
            if val is None or (isinstance(val, float) and val != val):
                cell.font = data_font_black
                cell.value = None
                continue

            cell.value = val
            if col_idx <= 2:
                cell.font = data_font_black
                cell.alignment = data_align_l
            else:
                cell.font = data_font_blue
                cell.alignment = data_align_r
                if fmt:
                    cell.number_format = fmt

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False


def _build_excel_export(df, mean_d, median_d, segment_name):
    """Build a single-segment Excel file (legacy, used by sub-segment view)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = segment_name or "Comps"
    _write_excel_sheet(ws, df, mean_d, median_d, segment_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_multi_segment_excel():
    """Build a multi-tab Excel workbook with all 7 healthcare segments."""
    from openpyxl import Workbook
    from config.settings import EXCEL_OVERRIDE_PATH
    from fetcher.db_manager import DBManager
    from config.settings import DB_PATH

    db = DBManager(DB_PATH)

    segments = [
        ("Pharma",                              "pharma"),
        ("Consumer Health",                     "consumer_health"),
        ("MedTech",                             "medtech"),
        ("Life Sci Tools",                      "life_sci_tools"),
        ("Asset-Light Services",                "services"),
        ("CDMOs",                               "cdmo"),
        ("Health Tech",                         "health_tech"),
    ]

    nrr_map, grr_map, _ = _load_nrr_data()

    overrides = _cached_load_overrides()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for tab_name, segment_key in segments:
        data = db.get_latest_snapshots(segment=segment_key)
        if overrides:
            data = apply_overrides(data, overrides, skip_sources={"factset"})
        if not data:
            continue

        df = build_comps_df(data)
        if df.empty:
            continue
        _inject_nrr(df, nrr_map or {}, grr_map or {})
        mean_d, median_d = compute_comps_summary(df)

        ws = wb.create_sheet(title=tab_name)
        _write_excel_sheet(ws, df, mean_d, median_d, tab_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Public entry point ────────────────────────────────────────────────────────

def render_comp_table(data, segment_name, show_sub_segments=False):
    """Render a Meritech-style comp table (Revenue view)."""
    if not data:
        st.warning(
            f"No data available for {segment_name}. "
            "Run the data fetcher to populate the database."
        )
        return

    # Inject sticky-header JS once per page load (idempotent via win.__ctStickyInit)
    _inject_sticky_js()

    # Download button at top
    try:
        xlsx_data = _build_multi_segment_excel()
        st.markdown(
            '<style>.stDownloadButton { margin-bottom: 0 !important; }'
            '.stDownloadButton button { font-size: 11px !important; padding: 2px 10px !important; }</style>',
            unsafe_allow_html=True,
        )
        st.download_button(
            label="Download Excel",
            data=xlsx_data,
            file_name="healthcare_multiples_comps.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        pass

    # Subtitle — units note + sort order
    st.markdown(
        '<p style="color:#94A3B8;font-size:11px;margin:0 0 2px 0;">'
        "All financials in millions ($mm). Sorted by TEV descending.</p>",
        unsafe_allow_html=True,
    )

    # Apply FactSet Excel overrides (silent — no UI notice)
    overrides = _cached_load_overrides()
    if overrides:
        data = apply_overrides(data, overrides, skip_sources={"factset"})

    nrr_map, grr_map, nrr_meta = _load_nrr_data()

    cfg = _CFG_REVENUE

    if show_sub_segments:
        _render_with_sub_segments(data, cfg, nrr_map, nrr_meta, segment_name, grr_map=grr_map)
    else:
        _render_simple_table(data, cfg, nrr_map, nrr_meta, segment_name, grr_map=grr_map)


def _render_simple_table(data, cfg, nrr_map=None, nrr_meta=None, segment_name="", grr_map=None):
    """Flat table — Mean/Median summary rows at top, all companies below."""
    df = build_comps_df(data)
    if df.empty:
        st.warning("No data to display.")
        return
    _inject_nrr(df, nrr_map or {}, grr_map or {})
    mean_d, median_d = compute_comps_summary(df)
    st.markdown(_build_table_html(df, mean_d, median_d, cfg), unsafe_allow_html=True)
    _render_ma_footnote(df)


def _render_with_sub_segments(data, cfg, nrr_map=None, nrr_meta=None, segment_name="", grr_map=None):
    """Table grouped by sub-segment, each group with its own summary rows."""
    # Discover sub-segments that actually appear in the data, preserving insertion order
    sub_order: list[str] = []
    for d in data:
        ss = d.get("sub_segment")
        if ss and ss not in sub_order:
            sub_order.append(ss)

    for sub_seg in sub_order:
        sub_data = [d for d in data if d.get("sub_segment") == sub_seg]
        if not sub_data:
            continue

        sub_name = SUB_SEGMENT_DISPLAY.get(sub_seg, sub_seg)
        st.markdown(f"#### {sub_name}")

        df = build_comps_df(sub_data)
        _inject_nrr(df, nrr_map or {}, grr_map or {})
        mean_d, median_d = compute_comps_summary(df)
        st.markdown(_build_table_html(df, mean_d, median_d, cfg), unsafe_allow_html=True)
        _render_ma_footnote(df)
        st.divider()

    # Full segment summary — summary rows only (no individual companies)
    st.markdown("#### Full Segment Summary")

    parts = [
        _CT_CSS,
        '<div class="ct-outer"><table class="ct-tbl">',
        _thead_html(cfg),
        "<tbody>",
        _summary_row_html(mean_all,   "Mean",   cfg, " ct-mean"),
        _summary_row_html(median_all, "Median", cfg, " ct-med"),
        "</tbody></table></div>",
    ]
    st.markdown("".join(parts), unsafe_allow_html=True)
    _render_ma_footnote(df_all)

