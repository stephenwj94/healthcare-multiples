#!/usr/bin/env python3
"""
Phase 5: Generate FactSet Override Excel Template
New file only — does NOT touch any existing data files.
"""
import json, sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

sys.path.insert(0, "/Users/jamesmaycher/Tech Multiples")
from config.company_registry import COMPANY_REGISTRY

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE   = PatternFill("solid", fgColor="1F3864")
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F0")
YELLOW      = PatternFill("solid", fgColor="FFFF00")
LIGHT_YELLOW= PatternFill("solid", fgColor="FFFACD")
LIGHT_GRAY  = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FONT_BLK= Font(name="Arial", size=10, bold=True, color="000000")
DATA_FONT   = Font(name="Arial", size=9)
BLUE_FONT   = Font(name="Arial", size=9, color="0000FF")
RED_FONT    = Font(name="Arial", size=9, color="FF0000")
GREEN_FONT  = Font(name="Arial", size=9, color="006100")
BOLD_FONT   = Font(name="Arial", size=9, bold=True)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left", vertical="center")
RIGHT       = Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text):
    cell.value = text
    cell.font = HDR_FONT
    cell.fill = DARK_BLUE
    cell.alignment = CENTER
    cell.border = thin_border()

# ── Load data ─────────────────────────────────────────────────────────────────
with open("missing_companies_classified.json") as f:
    missing_cos = json.load(f)

try:
    with open("completeness_report.json") as f:
        completeness = json.load(f)
    gap_companies = completeness.get("companies_with_critical_gaps", [])
except:
    gap_companies = []

# ── Schema fields (from DB) for empty columns ─────────────────────────────────
METRIC_COLS = [
    ("Market Cap ($M)",         "market_cap"),
    ("Enterprise Value ($M)",   "enterprise_value"),
    ("LTM Revenue ($M)",        "ltm_revenue"),
    ("NTM Revenue ($M)",        "ntm_revenue"),
    ("NTM Rev Growth %",        "ntm_revenue_growth"),
    ("Gross Margin %",          "gross_margin"),
    ("EBITDA Margin %",         "ebitda_margin"),
    ("NTM TEV/Rev (x)",         "ntm_tev_rev"),
    ("NTM TEV/EBITDA (x)",      "ntm_tev_ebitda"),
    ("LTM TEV/Rev (x)",         "ltm_tev_rev"),
    ("3Y Rev CAGR %",           "n3y_revenue_cagr"),
    ("Growth Adj Rev (x)",      "growth_adj_rev"),
    ("FCF Margin %",            "fcf_margin"),
    ("52-Wk High %",            "pct_52wk_high"),
    ("2W Price Chg %",          "price_change_2w"),
    ("2M Price Chg %",          "price_change_2m"),
    ("FY End Month",            "fy_end_month"),
]

CRITICAL_DB_FIELDS = {
    "ntm_tev_rev", "ntm_revenue_growth", "gross_margin",
    "ntm_revenue", "enterprise_value", "market_cap",
    "n3y_revenue_cagr",
}

SEGMENT_LABEL = {
    "pharma":          "Pharma",
    "consumer_health": "Consumer Health",
    "medtech":         "MedTech",
    "life_sci_tools":  "Life Sci Tools",
    "services":        "Services",
    "cdmo":            "CDMOs",
    "health_tech":     "Health Tech",
}

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 1 — New Companies to Add
# ════════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "New Companies to Add"

FIXED_COLS = ["Company Name", "Ticker (FactSet)", "Segment", "Priority",
              "Comp Set Source(s)", "Mkt Cap Approx", "FY End Mo"]
headers = FIXED_COLS + [m[0] for m in METRIC_COLS] + ["Notes"]

for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col_idx, value=h)
    apply_header(c, h)

ws1.row_dimensions[1].height = 36
ws1.freeze_panes = "A2"

# Sort: HIGH first, then MEDIUM, then LOW
priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
missing_sorted = sorted(missing_cos, key=lambda x: (priority_order.get(x.get("permira_priority","LOW"), 9), x["company"]))

for row_idx, co in enumerate(missing_sorted, 2):
    bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE_FILL
    
    # Fixed columns
    vals = [
        co["company"],
        f"{co['ticker']}-US",
        co["segment"],
        co.get("permira_priority", ""),
        ", ".join(co.get("appears_in", [])),
        co.get("market_cap_approx", ""),
        str(co.get("fy_end_month", "")),
    ]
    for col_idx, val in enumerate(vals, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.font = DATA_FONT
        c.fill = bg
        c.alignment = LEFT
        c.border = thin_border()
    
    # Metric columns — blue text on yellow = needs filling
    for col_idx in range(len(FIXED_COLS)+1, len(FIXED_COLS)+len(METRIC_COLS)+1):
        c = ws1.cell(row=row_idx, column=col_idx, value="")
        c.fill = LIGHT_YELLOW
        c.font = BLUE_FONT
        c.alignment = RIGHT
        c.border = thin_border()
    
    # Notes column
    notes_col = len(FIXED_COLS) + len(METRIC_COLS) + 1
    note = co.get("classification_reasoning", "")
    c = ws1.cell(row=row_idx, column=notes_col, value=note)
    c.font = DATA_FONT
    c.fill = bg
    c.alignment = LEFT
    c.border = thin_border()

# Auto-fit columns
col_widths = {1: 32, 2: 16, 3: 18, 4: 10, 5: 30, 6: 12, 7: 10}
for i, (label, _) in enumerate(METRIC_COLS, len(FIXED_COLS)+1):
    col_widths[i] = max(12, min(20, len(label) + 2))
col_widths[len(headers)] = 45  # Notes

for col_idx, width in col_widths.items():
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Data Gaps: Existing Companies
# ════════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Data Gaps — Existing Companies")

CRITICAL_DISPLAY = [m for m in METRIC_COLS if m[1] in CRITICAL_DB_FIELDS]
gap_headers = ["Company", "Ticker", "Segment", "Snapshot Date", "# Critical Missing"] + \
              [m[0] for m in CRITICAL_DISPLAY] + ["Notes"]

for col_idx, h in enumerate(gap_headers, 1):
    c = ws2.cell(row=1, column=col_idx, value=h)
    apply_header(c, h)

ws2.row_dimensions[1].height = 36
ws2.freeze_panes = "A2"

if gap_companies:
    for row_idx, cg in enumerate(gap_companies, 2):
        bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE_FILL
        
        # Find full registry entry
        seg_raw = cg.get("segment", "")
        seg_label = SEGMENT_LABEL.get(seg_raw, seg_raw)
        
        fixed_vals = [
            cg.get("name", cg["ticker"]),
            cg["ticker"],
            seg_label,
            cg.get("snapshot_date", ""),
            len(cg.get("missing_critical", [])),
        ]
        for col_idx, val in enumerate(fixed_vals, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            c.font = DATA_FONT
            c.fill = bg
            c.alignment = LEFT
            c.border = thin_border()
        
        # Critical metric columns
        for i, (label, db_field) in enumerate(CRITICAL_DISPLAY, len(fixed_vals)+1):
            is_miss = db_field in cg.get("missing_critical", [])
            c = ws2.cell(row=row_idx, column=i,
                         value="MISSING" if is_miss else "OK")
            c.font = RED_FONT if is_miss else GREEN_FONT
            c.fill = RED_FILL if is_miss else GREEN_FILL
            c.alignment = CENTER
            c.border = thin_border()
        
        # Notes
        notes_col = len(gap_headers)
        c = ws2.cell(row=row_idx, column=notes_col, value="Requires FactSet override data")
        c.font = DATA_FONT
        c.fill = bg
        c.alignment = LEFT
        c.border = thin_border()
else:
    # DB empty or no gaps — add note
    c = ws2.cell(row=2, column=1, value="Database is empty or no critical gaps found — run fetcher first.")
    c.font = DATA_FONT
    
    # Still populate with registry companies as placeholders
    from config.company_registry import COMPANY_REGISTRY
    for row_idx, co in enumerate(COMPANY_REGISTRY, 3):
        bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE_FILL
        seg_label = SEGMENT_LABEL.get(co["segment"], co["segment"])
        fixed_vals = [co["name"], co["ticker"], seg_label, "Not yet fetched", "Unknown"]
        for col_idx, val in enumerate(fixed_vals, 1):
            c2 = ws2.cell(row=row_idx, column=col_idx, value=val)
            c2.font = DATA_FONT; c2.fill = bg; c2.alignment = LEFT; c2.border = thin_border()
        for i, (label, db_field) in enumerate(CRITICAL_DISPLAY, len(fixed_vals)+1):
            c2 = ws2.cell(row=row_idx, column=i, value="UNKNOWN")
            c2.font = Font(name="Arial", size=9, color="7F7F7F")
            c2.fill = LIGHT_YELLOW; c2.alignment = CENTER; c2.border = thin_border()

for col_idx in range(1, len(gap_headers)+1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 16
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 18

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Full Company Master
# ════════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Full Company Master")

master_headers = ["Company", "Ticker (FactSet)", "Segment", "Sub-Segment",
                  "Status", "FY End Mo", "Reporting Currency"] + \
                 [m[0] for m in METRIC_COLS] + ["Notes"]

for col_idx, h in enumerate(master_headers, 1):
    c = ws3.cell(row=1, column=col_idx, value=h)
    apply_header(c, h)
ws3.row_dimensions[1].height = 36
ws3.freeze_panes = "A2"

# All existing companies first (sorted by segment then name)
all_master = []
for co in COMPANY_REGISTRY:
    all_master.append({
        "company": co["name"],
        "ticker": co["ticker"],
        "segment": SEGMENT_LABEL.get(co["segment"], co["segment"]),
        "sub_segment": co.get("sub_segment", "") or "",
        "status": "EXISTING",
        "fy_end_month": co.get("fy_end_month", ""),
        "currency": co.get("reporting_currency", "USD"),
    })
for co in missing_sorted:
    all_master.append({
        "company": co["company"],
        "ticker": co["ticker"],
        "segment": co["segment"],
        "sub_segment": "",
        "status": "NEW — NEEDS DATA",
        "fy_end_month": co.get("fy_end_month", ""),
        "currency": "USD",
    })

all_master.sort(key=lambda x: (x["segment"], x["company"]))

for row_idx, co in enumerate(all_master, 2):
    is_new = co["status"] == "NEW — NEEDS DATA"
    bg = LIGHT_YELLOW if is_new else (LIGHT_GRAY if row_idx % 2 == 0 else WHITE_FILL)
    
    fixed_vals = [
        co["company"], f"{co['ticker']}-US" if is_new else co["ticker"],
        co["segment"], co["sub_segment"], co["status"],
        co["fy_end_month"], co["currency"],
    ]
    for col_idx, val in enumerate(fixed_vals, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.font = BLUE_FONT if is_new else DATA_FONT
        c.fill = bg
        c.alignment = LEFT
        c.border = thin_border()
    
    for i in range(len(fixed_vals)+1, len(fixed_vals)+len(METRIC_COLS)+1):
        c = ws3.cell(row=row_idx, column=i, value="" if is_new else "See DB")
        c.fill = LIGHT_YELLOW if is_new else bg
        c.font = BLUE_FONT if is_new else Font(name="Arial", size=9, color="808080")
        c.alignment = RIGHT
        c.border = thin_border()
    
    c = ws3.cell(row=row_idx, column=len(master_headers),
                 value="Needs FactSet override" if is_new else "")
    c.font = DATA_FONT; c.fill = bg; c.alignment = LEFT; c.border = thin_border()

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Classification Reference
# ════════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Classification Reference")

def ref_section(ws, start_row, title, definition, examples, focus):
    c = ws.cell(row=start_row, column=1, value=title)
    c.font = Font(name="Arial", size=12, bold=True, color="1F3864")
    c.fill = LIGHT_BLUE
    c.alignment = LEFT
    ws.merge_cells(f"A{start_row}:F{start_row}")
    
    c = ws.cell(row=start_row+1, column=1, value="Definition")
    c.font = Font(name="Arial", size=9, bold=True)
    c = ws.cell(row=start_row+1, column=2, value=definition)
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{start_row+1}:F{start_row+1}")
    ws.row_dimensions[start_row+1].height = 40
    
    c = ws.cell(row=start_row+2, column=1, value="Examples")
    c.font = Font(name="Arial", size=9, bold=True)
    c = ws.cell(row=start_row+2, column=2, value=examples)
    c.font = Font(name="Arial", size=9)
    ws.merge_cells(f"B{start_row+2}:F{start_row+2}")
    
    c = ws.cell(row=start_row+3, column=1, value="Permira Focus")
    c.font = Font(name="Arial", size=9, bold=True)
    c = ws.cell(row=start_row+3, column=2, value=focus)
    c.font = Font(name="Arial", size=9, color="1F3864")
    ws.merge_cells(f"B{start_row+3}:F{start_row+3}")
    return start_row + 6

ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 80

r = 1
c = ws4.cell(row=r, column=1, value="Permira Segment Classification Reference")
c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
c.fill = DARK_BLUE
c.alignment = CENTER
ws4.merge_cells(f"A{r}:F{r}")
ws4.row_dimensions[r].height = 28
r += 2

r = ref_section(ws4, r, "Pharma",
    "Branded pharmaceutical and biotech companies. Patent-protected revenue, R&D-heavy, "
    "regulatory and clinical milestones drive value.",
    "Eli Lilly, Pfizer, Merck, Bristol Myers Squibb, AbbVie, Novartis, AstraZeneca, GSK, Gilead",
    "Focus on innovative pipelines, late-stage clinical readouts, patent cliffs, and acquisitions of mid-cap biotech."
)
r = ref_section(ws4, r, "Consumer Health",
    "OTC, vitamins, personal care, and consumer-facing wellness brands.",
    "Haleon, Kenvue, Procter & Gamble (Health), Reckitt, Church & Dwight",
    "Focus on share trends, brand strength, gross margin resilience, and pricing power."
)
r = ref_section(ws4, r, "MedTech",
    "Medical devices, diagnostics, and equipment. Capital equipment + recurring consumables.",
    "Medtronic, Stryker, Boston Scientific, Abbott, Becton Dickinson, Edwards Lifesciences, Intuitive Surgical, "
    "Zimmer Biomet, Hologic, Dexcom",
    "Focus on procedure volume trends, new product cycles, and category leadership."
)
r = ref_section(ws4, r, "Life Sci Tools / Dx / Bioprocessing",
    "Tools, instruments, reagents, and bioprocessing equipment for life-sciences research and biomanufacturing.",
    "Thermo Fisher, Danaher, Agilent, Waters, Bio-Rad, Bio-Techne, Sartorius, Repligen, Avantor, Revvity",
    "Focus on funding environment, biopharma capex cycle, and bioprocessing capacity utilization."
)
r = ref_section(ws4, r, "Asset-Light Services",
    "Healthcare services with low capital intensity — staffing, distribution, GPOs, IT services, payer services.",
    "McKesson, Cencora, Cardinal Health, Premier, R1 RCM, Evolent Health",
    "Focus on margin trajectory, contract wins, and capital-light recurring revenue."
)
r = ref_section(ws4, r, "CDMOs",
    "Contract development and manufacturing organizations serving biopharma — small molecule, biologics, cell & gene therapy.",
    "Catalent, Lonza, Samsung Biologics, WuXi Biologics, Charles River, IQVIA",
    "Focus on capacity utilization, biologics mix, and biotech funding signals."
)
r = ref_section(ws4, r, "Health Tech",
    "Healthcare-focused software, telehealth, RCM tech, clinical workflow, and digital therapeutics.",
    "Veeva, Doximity, Hims & Hers, Teladoc, GoodRx, Schrodinger, Waystar, Phreesia",
    "Focus on attach rates, customer concentration, and AI-driven workflow automation."
)

# ── Save ──────────────────────────────────────────────────────────────────────
today_str = date.today().strftime("%Y%m%d")
out_path = f"/Users/jamesmaycher/Tech Multiples/data/FactSet_Override_Template_{today_str}.xlsx"
wb.save(out_path)
print(f"[Phase 5] Excel saved: {out_path}")
print(f"  Sheet 1 — New Companies to Add: {len(missing_sorted)} rows")
print(f"  Sheet 2 — Data Gaps: {max(len(gap_companies), len(COMPANY_REGISTRY))} rows")
print(f"  Sheet 3 — Full Company Master: {len(all_master)} rows")
print(f"  Sheet 4 — Classification Reference: written")
